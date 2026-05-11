# futgg_evo_crawler_fixed_v6.py
# FUT.GG Current Evolutions + Rewards -> unified Excel crawler
#
# v8
# - Avoid long networkidle timeouts on FUT.GG detail pages.
# - Use domcontentloaded + short load wait, and fail bad runs to avoid overwriting latest files.
# - Based on the early version whose stat parsing worked.
# - Only the URL collection/source classification part was changed.
# - Current and Reward links are collected separately from their own tab states.
# - Detail parsing logic is intentionally kept close to the early working version.
#
# Run locally:
#   py futgg_evolution_crawler.py --show
#   py futgg_evolution_crawler.py --out "D:\\FUTGG"
#
# Run on GitHub Actions:
#   python futgg_evolution_crawler.py --out outputs --headless
#
# Output:
#   By default, output files are saved in the same folder as this .py file.
#   Use --out to choose another output folder.

from __future__ import annotations

import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.fut.gg"
START_URL = "https://www.fut.gg/evolutions/"

# Portable path settings
# - Default output folder: the same folder as this .py file
# - Override output folder:
#     --out "D:\\FUTGG"
# - For GitHub Actions:
#     --out outputs --headless
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUT_DIR = SCRIPT_DIR

# Existing analysis workbook update is optional.
# GitHub Actions should normally NOT update an analysis workbook.
# Local example:
#   py futgg_evolution_crawler.py --show --update-analysis --analysis "D:\\FUTGG\\futgg_evolutions_analysis.xlsx"
DEFAULT_ANALYSIS_FILE_NAME = "futgg_evolutions_analysis.xlsx"

# True: 만료(Expired) 표시가 있는 Evolution은 제외
EXCLUDE_EXPIRED = True

# FUT.GG often keeps background network requests open.
# Waiting for "networkidle" can make every detail page hang until timeout.
LIST_GOTO_TIMEOUT_MS = 30000
DETAIL_GOTO_TIMEOUT_MS = 20000
POST_GOTO_LOAD_WAIT_MS = 5000
MIN_SUCCESS_ROWS = 5
MIN_SUCCESS_RATE = 0.20

FACE_STATS = ["OVR", "PAC", "SHO", "PAS", "DRI", "DEF", "PHY"]
SUB_STATS = [
    "Acceleration", "Sprint Speed",
    "Positioning", "Finishing", "Shot Power", "Long Shots", "Volleys", "Penalties",
    "Vision", "Crossing", "FK Accuracy", "Short Passing", "Long Passing", "Curve",
    "Agility", "Balance", "Reactions", "Ball Control", "Dribbling", "Composure",
    "Interceptions", "Heading Accuracy", "Defensive Awareness", "Standing Tackle", "Sliding Tackle",
    "Jumping", "Stamina", "Strength", "Aggression",
]
SPECIAL_STATS = ["SM", "WF"]
ALL_STATS = FACE_STATS + SUB_STATS + SPECIAL_STATS

STAT_ALIASES = {
    "Att. Position": "Positioning",
    "Positioning": "Positioning",
    "FK Acc.": "FK Accuracy",
    "FK Accuracy": "FK Accuracy",
    "Short Pass": "Short Passing",
    "Short Passing": "Short Passing",
    "Long Pass": "Long Passing",
    "Long Passing": "Long Passing",
    "Heading Acc.": "Heading Accuracy",
    "Heading Accuracy": "Heading Accuracy",
    "Def. Aware": "Defensive Awareness",
    "Defensive Awareness": "Defensive Awareness",
    "Stand Tackle": "Standing Tackle",
    "Standing Tackle": "Standing Tackle",
    "Slide Tackle": "Sliding Tackle",
    "Sliding Tackle": "Sliding Tackle",
}

STAT_PATTERNS = sorted(
    set(FACE_STATS + SPECIAL_STATS + SUB_STATS + list(STAT_ALIASES.keys())),
    key=len,
    reverse=True,
)
STAT_PATTERN = "|".join(re.escape(s) for s in STAT_PATTERNS)

REQ_PATTERNS = [
    ("Req_OVR_Max", r"Overall Max\.?\s*([0-9]+)"),
    ("Req_OVR_Min", r"Overall Min\.?\s*([0-9]+)"),
    ("Req_PAC_Max", r"Pace Max\.?\s*([0-9]+)"),
    ("Req_SHO_Max", r"Shooting Max\.?\s*([0-9]+)"),
    ("Req_PAS_Max", r"Passing Max\.?\s*([0-9]+)"),
    ("Req_DRI_Max", r"Dribbling Max\.?\s*([0-9]+)"),
    ("Req_DEF_Max", r"Defending Max\.?\s*([0-9]+)"),
    ("Req_PHY_Max", r"Physicality Max\.?\s*([0-9]+)"),
    ("Req_Max_PS_Plus", r"Max PS\+\s*([0-9]+)"),
    ("Req_Max_PS", r"Max PS\s*([0-9]+)"),
    ("Req_Max_Positions", r"Max Pos\.?\s*([0-9]+)"),
    ("Req_Born_Before", r"Born Before\s*([0-9]{4}-[0-9]{2}-[0-9]{2})"),
    ("Req_Born_After", r"Born After\s*([0-9]{4}-[0-9]{2}-[0-9]{2})"),
]


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def normalize_stat(name: str) -> str:
    name = clean_text(name)
    return STAT_ALIASES.get(name, name)


def unique_keep_order(items):
    seen = set()
    out = []
    for x in items:
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def canonical_url(url: str) -> str:
    return (url or "").split("?", 1)[0].rstrip("/") + "/"


def is_detail_evolution_url(url: str) -> bool:
    # Keep only true detail pages, e.g. /evolutions/1962-rapid-gains/
    return re.search(r"/evolutions/[0-9]+-[^/]+/?$", canonical_url(url)) is not None


def unique_rows_by_url(rows: List[Dict]) -> List[Dict]:
    seen = set()
    out = []
    for r in rows:
        u = r.get("URL")
        if not u or u in seen:
            continue
        seen.add(u)
        out.append(r)
    return out


def infer_reward_category(text: str) -> str:
    t = clean_text(text)
    if re.search(r"Premium Season Pass|SP\+", t, re.I):
        return "Premium Season Pass"
    if re.search(r"Standard Season Pass|\[SP\b|\bSP\s*\d+", t, re.I):
        return "Standard Season Pass"
    if re.search(r"\bSBC\b", t, re.I):
        return "SBC"
    if re.search(r"Objective", t, re.I):
        return "Objective"
    return "Reward"


def is_expired_text(text: str) -> bool:
    """
    FUT.GG 목록/상세페이지에서 'Expired'로 표시된 진화 제외용.
    'Expires'는 아직 만료 전 남은 시간 표시이므로 제외하지 않습니다.
    """
    t = clean_text(text)
    return re.search(r"\bExpired\b", t, re.I) is not None


def safe_goto(page, url: str, timeout: int, label: str = "page"):
    """
    Navigate without waiting for networkidle.
    FUT.GG pages can keep analytics/image/API requests open, so networkidle may
    never arrive even when the useful DOM text is already available.
    """
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=timeout)
        try:
            page.wait_for_load_state("load", timeout=POST_GOTO_LOAD_WAIT_MS)
        except Exception:
            pass
        time.sleep(0.8)
        return True, None
    except Exception as e:
        # Sometimes useful body text exists even after a navigation timeout.
        try:
            body = page.inner_text("body", timeout=3000)
            if clean_text(body):
                print(f"  WARN: {label} goto timed out, but body text exists. Continue parsing.")
                return True, None
        except Exception:
            pass
        return False, e


def click_tab_contains(page, label: str) -> bool:
    pattern = re.compile(label, re.I)
    candidates = [
        page.get_by_role("tab", name=pattern),
        page.get_by_role("button", name=pattern),
        page.get_by_role("link", name=pattern),
        page.locator("button").filter(has_text=pattern),
        page.locator("a").filter(has_text=pattern),
        page.locator("[role='tab']").filter(has_text=pattern),
        page.get_by_text(pattern),
    ]

    for loc in candidates:
        try:
            count = loc.count()
            for i in range(min(count, 8)):
                target = loc.nth(i)
                if target.is_visible(timeout=1000):
                    target.scroll_into_view_if_needed(timeout=2000)
                    target.click(timeout=5000, force=True)
                    try:
                        page.wait_for_load_state("load", timeout=5000)
                    except Exception:
                        pass
                    time.sleep(1.5)
                    return True
        except Exception:
            continue

    return False


def click_show_all_until_done(page, max_clicks: int = 20):
    # This is intentionally the early working version style.
    for _ in range(max_clicks):
        try:
            loc = page.get_by_text(re.compile(r"Show all", re.I))
            n = loc.count()
            if n == 0:
                return
            clicked = False
            for i in range(min(n, 5)):
                try:
                    loc.nth(i).click(timeout=1000)
                    clicked = True
                    time.sleep(0.2)
                except Exception:
                    pass
            if not clicked:
                return
        except Exception:
            return


def collect_evolution_link_rows(page, source_group: str) -> List[Dict]:
    """
    Early-version style collection:
    - Read all anchors from the currently selected page/tab.
    - Filter to true detail URLs only.
    - Do not use viewport-only logic because it missed many cards.
    """
    click_show_all_until_done(page, max_clicks=5)

    rows = page.eval_on_selector_all(
        "a[href]",
        """
        els => els.map(a => {
            let parent = a;
            for (let i = 0; i < 5; i++) {
                if (!parent.parentElement) break;
                parent = parent.parentElement;
                const txt = (parent.innerText || parent.textContent || '').trim();
                if (txt.length > 30) break;
            }
            return {
                href: a.href,
                link_text: (a.innerText || a.textContent || '').trim(),
                parent_text: (parent.innerText || parent.textContent || '').trim()
            };
        })
        """
    )

    out = []
    for r in rows:
        url = canonical_url(r.get("href", ""))
        if not is_detail_evolution_url(url):
            continue

        link_text = clean_text(r.get("link_text") or "")
        parent_text = clean_text(r.get("parent_text") or "")

        # 목록 카드에 Expired 표시가 있으면 수집 제외
        if EXCLUDE_EXPIRED and is_expired_text(parent_text + " " + link_text):
            continue

        category = "Current Evolutions" if source_group == "Current" else infer_reward_category(parent_text + " " + link_text)

        out.append({
            "Source_Group": source_group,
            "Category": category,
            "URL": url,
            "Link_Text": link_text,
            "Parent_Text": parent_text,
        })

    return unique_rows_by_url(out)


def get_all_link_rows(headless: bool = True) -> List[Dict]:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)

        # Current: collect from a fresh default Evolutions page. Do not click generic "Evolutions".
        page_current = browser.new_page(viewport={"width": 1400, "height": 1200})
        ok, err = safe_goto(page_current, START_URL, timeout=LIST_GOTO_TIMEOUT_MS, label="Current list")
        if not ok:
            raise RuntimeError(f"Current list page load failed: {type(err).__name__}: {err}")
        time.sleep(1)

        print("[1/2] Collecting Current Evolutions links...")
        current_rows = collect_evolution_link_rows(page_current, "Current")
        print(f"  - Current detail links: {len(current_rows)}")
        page_current.close()

        # Reward: open fresh page, click Rewards, then collect all detail links in that selected tab.
        page_reward = browser.new_page(viewport={"width": 1400, "height": 1200})
        ok, err = safe_goto(page_reward, START_URL, timeout=LIST_GOTO_TIMEOUT_MS, label="Reward list")
        if not ok:
            raise RuntimeError(f"Reward list page load failed: {type(err).__name__}: {err}")
        time.sleep(1)

        print("[2/2] Collecting Rewards links...")
        if click_tab_contains(page_reward, "Rewards"):
            reward_rows = collect_evolution_link_rows(page_reward, "Reward")
            print(f"  - Reward detail links: {len(reward_rows)}")
        else:
            reward_rows = []
            print("  - Rewards tab click not detected. Try running with --show.")
        page_reward.close()

        browser.close()

    # Keep Current first. If a URL appears in both, Current wins.
    current_rows = unique_rows_by_url(current_rows)
    reward_rows = unique_rows_by_url(reward_rows)
    current_urls = {r["URL"] for r in current_rows}
    combined = current_rows + [r for r in reward_rows if r["URL"] not in current_urls]

    print(f"Collected {len(combined)} unique detail URLs.")
    print(f"  Current kept: {sum(1 for r in combined if r['Source_Group'] == 'Current')}")
    print(f"  Reward kept: {sum(1 for r in combined if r['Source_Group'] == 'Reward')}")
    return combined


def extract_between(text: str, start_label: str, end_labels: list[str]) -> str:
    s = clean_text(text)
    start = re.search(re.escape(start_label), s, re.I)
    if not start:
        return ""
    start_idx = start.end()
    end_idx = len(s)
    for lab in end_labels:
        m = re.search(re.escape(lab), s[start_idx:], re.I)
        if m:
            end_idx = min(end_idx, start_idx + m.start())
    return clean_text(s[start_idx:end_idx])


def parse_costs(text: str):
    s = clean_text(text)
    coins = fp = None
    m = re.search(r"(?<![0-9])([0-9]{1,3}(?:,[0-9]{3})+)\s+([0-9]{2,4})(?![0-9])", s)
    if m:
        try:
            coins = int(m.group(1).replace(",", ""))
            fp = int(m.group(2))
        except Exception:
            pass
    return coins, fp


def parse_requirements(text: str):
    s = clean_text(text)
    req_text = extract_between(s, "Requirements", ["Upgrades", "Challenges", "Training Time"])
    result = {f"Req_{k}": None for k in [
        "OVR_Max", "OVR_Min", "PAC_Max", "SHO_Max", "PAS_Max", "DRI_Max", "DEF_Max", "PHY_Max",
        "Position", "Excluded_Position", "Rarity", "Excluded_Rarity",
        "Max_PS", "Max_PS_Plus", "Max_Positions", "Born_Before", "Born_After", "Other"
    ]}
    long_rows = []

    if not req_text:
        result["Req_Other"] = ""
        return result, long_rows, req_text

    for col, pat in REQ_PATTERNS:
        m = re.search(pat, req_text, re.I)
        if m:
            result[col] = m.group(1)
            long_rows.append({
                "Requirement_Category": col.replace("Req_", ""),
                "Requirement_Name": col.replace("Req_", ""),
                "Operator": "Max" if col.endswith("_Max") or "Max" in col else "Value",
                "Value": m.group(1),
                "Raw_Text": m.group(0),
            })

    def cap_list(label: str, next_words: list[str]):
        pat = re.escape(label) + r"\s+(.+?)(?=\s+(?:" + "|".join(re.escape(w) for w in next_words) + r")\b|$)"
        m = re.search(pat, req_text, re.I)
        if not m:
            return None, None
        val = clean_text(m.group(1))
        return val, m.group(0)

    next_req_labels = [
        "Excluded Rarity", "Rarity", "Overall", "Position", "Excluded Position",
        "Pace", "Shooting", "Passing", "Dribbling", "Defending", "Physicality",
        "Born Before", "Born After", "Max PS+", "Max PS", "Max Pos.", "Upgrades"
    ]

    for label, col in [
        ("Excluded Position", "Req_Excluded_Position"),
        ("Position", "Req_Position"),
        ("Excluded Rarity", "Req_Excluded_Rarity"),
        ("Rarity", "Req_Rarity"),
    ]:
        val, raw = cap_list(label, [x for x in next_req_labels if x != label])
        if val:
            result[col] = val
            long_rows.append({
                "Requirement_Category": col.replace("Req_", ""),
                "Requirement_Name": label,
                "Operator": "Include" if "Excluded" not in label else "Exclude",
                "Value": val,
                "Raw_Text": raw,
            })

    result["Req_Other"] = req_text
    return result, long_rows, req_text


def parse_upgrades(text: str):
    s = clean_text(text)
    upgrade_text = extract_between(s, "Upgrades", ["Challenges", "Training Time", "Objectives", "Expires", "©"])
    if not upgrade_text:
        upgrade_text = s

    rows = []
    pat = re.compile(
        rf"\+\s*([0-9]+)\s+({STAT_PATTERN})(?:\s+([0-9]{{1,3}}))?",
        re.I
    )
    for m in pat.finditer(upgrade_text):
        inc = int(m.group(1))
        stat = normalize_stat(m.group(2))
        cap = m.group(3)
        if stat in FACE_STATS:
            utype = "Face Stat"
        elif stat in SPECIAL_STATS:
            utype = "Skill"
        else:
            utype = "Sub Stat"
        rows.append({
            "Level": None,
            "Upgrade_Type": utype,
            "Stat_Name": stat,
            "Increase": inc,
            "Cap_or_Result": int(cap) if cap and cap.isdigit() else None,
            "Raw_Text": clean_text(m.group(0)),
        })

    role_bits = []
    role_pat = re.compile(r"\b(GK|LB|LWB|CB|RB|RWB|CDM|CM|CAM|LM|RM|LW|RW|ST)\s+[A-Za-z][A-Za-z ]+\s*\+{1,2}")
    for m in role_pat.finditer(upgrade_text):
        role_bits.append(clean_text(m.group(0)))

    return rows, unique_keep_order(role_bits), upgrade_text


def matrix_base_row():
    row = {}
    for stat in ALL_STATS:
        row[f"+ {stat}"] = None
        row[f"Max {stat}"] = None
    return row


def merge_upgrade_to_matrix(row, upgrade_rows):
    for u in upgrade_rows:
        stat = u["Stat_Name"]
        if stat not in ALL_STATS:
            continue
        plus_col = f"+ {stat}"
        max_col = f"Max {stat}"
        inc = u.get("Increase")
        cap = u.get("Cap_or_Result")

        if inc is not None:
            row[plus_col] = (row.get(plus_col) or 0) + inc
        if cap is not None:
            prev = row.get(max_col)
            row[max_col] = max(prev or 0, cap)
    return row


def scrape_detail(context, link_row: Dict):
    # This is intentionally close to the early working version.
    url = link_row["URL"]
    page = context.new_page()
    try:
        ok, nav_err = safe_goto(page, url, timeout=DETAIL_GOTO_TIMEOUT_MS, label="detail")
        if not ok:
            return None, [], [], f"{url} :: {type(nav_err).__name__}: {nav_err}"

        click_show_all_until_done(page, max_clicks=20)
        body = page.inner_text("body", timeout=15000)

        # 상세페이지에서도 Expired 표시가 있으면 최종 제외
        # 목록에서 못 걸러진 경우를 대비한 안전장치입니다.
        if EXCLUDE_EXPIRED and is_expired_text(body):
            print(f"  SKIP expired: {url}")
            return None, [], [], None

        title = None
        try:
            h1s = page.locator("h1").all_inner_texts()
            if h1s:
                title = clean_text(h1s[0])
        except Exception:
            pass
        if not title:
            slug = url.rstrip("/").split("/")[-1]
            title = re.sub(r"^\d+-", "", slug).replace("-", " ").title()

        coins, fp = parse_costs(body)
        req_matrix, req_long, req_text = parse_requirements(body)
        up_rows, roles, up_text = parse_upgrades(body)

        matrix = {
            "Source_Group": link_row.get("Source_Group"),
            "Category": link_row.get("Category"),
            "Evolution_Name": title,
            "URL": url,
            "Cost_Coins": coins,
            "Cost_FP": fp,
            "Raw_Requirements_Text": req_text,
            "Role_Added": "; ".join(roles) if roles else None,
            "Other_Upgrade_Text": up_text,
            "Link_Text": link_row.get("Link_Text"),
            "Parent_Text": link_row.get("Parent_Text"),
        }
        matrix.update(req_matrix)
        matrix.update(matrix_base_row())
        matrix = merge_upgrade_to_matrix(matrix, up_rows)

        for r in req_long:
            r.update({
                "Evolution_Name": title,
                "Source_Group": link_row.get("Source_Group"),
                "Category": link_row.get("Category"),
                "URL": url,
            })

        for u in up_rows:
            u.update({
                "Evolution_Name": title,
                "Source_Group": link_row.get("Source_Group"),
                "Category": link_row.get("Category"),
                "URL": url,
            })

        return matrix, req_long, up_rows, None
    except Exception as e:
        return None, [], [], f"{url} :: {type(e).__name__}: {e}"
    finally:
        page.close()





def update_status_sheet_layout(wb, df_matrix):
    """(현황표) 시트를 상위 3개 순위 요약표(4행 블록 x 3개)로 채웁니다."""
    if "(현황표)" not in wb.sheetnames:
        return
    ws = wb["(현황표)"]
    start_col, end_col = 11, 46  # K:AT

    stat_point_col = "+ Stat Point" if "+ Stat Point" in df_matrix.columns else None
    plus_cols = [c for c in df_matrix.columns if isinstance(c, str) and c.startswith("+ ")]
    if "+ OVR" in plus_cols:
        plus_cols.remove("+ OVR")

    # 기존 병합 해제 후 1~3위 병합 재설정
    for rng in list(ws.merged_cells.ranges):
        if rng.min_col == 8 and rng.max_col == 8 and rng.min_row <= 13 and rng.max_row >= 2:
            ws.unmerge_cells(str(rng))
    for base_row, rank_name in ((2, "1위"), (6, "2위"), (10, "3위")):
        ws.merge_cells(start_row=base_row, start_column=8, end_row=base_row + 3, end_column=8)
        ws.cell(row=base_row, column=8, value=rank_name)

    for rank in (1, 2, 3):
        base_row = 2 + (rank - 1) * 4
        ws.cell(row=base_row, column=9, value="MAX")
        ws.cell(row=base_row + 1, column=9, value="+OVR")
        ws.cell(row=base_row + 2, column=9, value="+Stat Point")
        ws.cell(row=base_row + 3, column=9, value="진화명")
        ws.row_dimensions[base_row].height = 20
        ws.row_dimensions[base_row + 1].height = 20
        ws.row_dimensions[base_row + 2].height = 20
        ws.row_dimensions[base_row + 3].height = 56

        for c in range(start_col, end_col + 1):
            stat_name = ws.cell(row=1, column=c).value
            max_col = f"Max {stat_name}" if stat_name is not None else None
            if not max_col or max_col not in df_matrix.columns:
                ws.cell(row=base_row, column=c, value=None)
                ws.cell(row=base_row + 1, column=c, value=None)
                ws.cell(row=base_row + 2, column=c, value=None)
                ws.cell(row=base_row + 3, column=c, value=None)
                continue

            series = pd.to_numeric(df_matrix[max_col], errors="coerce")
            valid = df_matrix.loc[series.notna() & (series > 0)].copy()
            if valid.empty:
                ws.cell(row=base_row, column=c, value=None)
                ws.cell(row=base_row + 1, column=c, value=None)
                ws.cell(row=base_row + 2, column=c, value=None)
                ws.cell(row=base_row + 3, column=c, value=None)
                continue

            ranked_vals = sorted(valid[max_col].dropna().unique(), reverse=True)
            if len(ranked_vals) < rank:
                ws.cell(row=base_row, column=c, value=None)
                ws.cell(row=base_row + 1, column=c, value=None)
                ws.cell(row=base_row + 2, column=c, value=None)
                ws.cell(row=base_row + 3, column=c, value=None)
                continue

            target_max = ranked_vals[rank - 1]
            row = valid.loc[valid[max_col] == target_max].iloc[0]
            ovr_val = row.get("+ OVR", None)
            if stat_point_col:
                stat_point_val = row.get(stat_point_col, None)
            else:
                stat_point_val = pd.to_numeric(pd.Series([row.get(pc, 0) for pc in plus_cols]), errors="coerce").fillna(0).sum()
            evo_name = row.get("Evolution_Name", None)

            ws.cell(row=base_row, column=c, value=target_max)
            ws.cell(row=base_row + 1, column=c, value=None if pd.isna(ovr_val) else ovr_val)
            ws.cell(row=base_row + 2, column=c, value=None if pd.isna(stat_point_val) else stat_point_val)
            ws.cell(row=base_row + 3, column=c, value=None if pd.isna(evo_name) else str(evo_name))
            ws.cell(row=base_row + 3, column=c).alignment = ws.cell(row=base_row + 3, column=c).alignment.copy(wrap_text=True)

def update_analysis_workbook(df_matrix, analysis_path):
    """
    기존 분석용 엑셀 파일의 Unified_Matrix 시트에 있는 Evo 표 데이터를
    새 크롤링 결과(df_matrix)로 교체하고, 표 범위를 새 행 수에 맞게 갱신합니다.

    주의:
    - analysis_path 파일은 Excel에서 닫혀 있어야 합니다.
    - Unified_Matrix 시트의 1행 머리글과 df_matrix 열 이름이 일치해야 합니다.
    - Evo 표 이름이 유지되어 있어야 합니다.
    """
    wb = load_workbook(analysis_path)
    update_status_sheet_layout(wb, df_matrix)
    ws = wb["Unified_Matrix"]

    if "Evo" not in ws.tables:
        raise ValueError("Unified_Matrix 시트에서 Evo 표를 찾을 수 없습니다.")

    table = ws.tables["Evo"]

    # 기존 분석파일의 머리글 순서를 기준으로 새 데이터 열 순서를 맞춥니다.
    headers = []
    col = 1
    while ws.cell(row=1, column=col).value is not None:
        headers.append(ws.cell(row=1, column=col).value)
        col += 1

    missing_cols = [c for c in headers if c not in df_matrix.columns]
    if missing_cols:
        raise ValueError(f"크롤링 결과에 없는 열이 있습니다: {missing_cols}")

    df_out = df_matrix[headers].copy()

    max_col = len(headers)

    # 기존 표 데이터 행 삭제. 머리글 1행은 유지합니다.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # 새 데이터 입력
    for r_idx, row in enumerate(df_out.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            # pandas NaN은 빈칸으로 입력
            if pd.isna(value):
                value = None
            ws.cell(row=r_idx, column=c_idx, value=value)

    last_row = len(df_out) + 1
    last_col_letter = get_column_letter(max_col)

    # Evo 표 범위 재설정
    table.ref = f"A1:{last_col_letter}{last_row}"

    wb.save(analysis_path)



def get_cli_value(flag: str) -> Optional[str]:
    if flag not in sys.argv:
        return None
    idx = sys.argv.index(flag)
    if idx + 1 >= len(sys.argv):
        return None
    return sys.argv[idx + 1]


def resolve_out_dir() -> Path:
    """
    Output folder priority:
      1) --out "folder path"
      2) same folder as this script
    """
    out_arg = get_cli_value("--out")
    if out_arg:
        return Path(out_arg).expanduser().resolve()
    return DEFAULT_OUT_DIR


def resolve_headless(default_headless: bool = True) -> bool:
    """
    Browser visibility priority:
      1) --show => headless False
      2) --headless => headless True
      3) default
    """
    if "--show" in sys.argv:
        return False
    if "--headless" in sys.argv:
        return True
    return default_headless


def resolve_update_analysis() -> bool:
    """
    Analysis workbook update is OFF by default.
    Turn on only when running locally with:
      --update-analysis
    """
    return "--update-analysis" in sys.argv


def resolve_analysis_file(out_dir: Path) -> Path:
    """
    Analysis workbook path priority:
      1) --analysis "path/to/file.xlsx"
      2) out_dir / DEFAULT_ANALYSIS_FILE_NAME
    """
    analysis_arg = get_cli_value("--analysis")
    if analysis_arg:
        return Path(analysis_arg).expanduser().resolve()
    return out_dir / DEFAULT_ANALYSIS_FILE_NAME


def main():
    today = datetime.now().strftime("%Y%m%d")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    out_dir = resolve_out_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"futgg_evolutions_unified_{timestamp}.xlsx"

    headless = resolve_headless(default_headless=True)
    update_analysis = resolve_update_analysis()
    analysis_file = resolve_analysis_file(out_dir)

    print("FUT.GG Evolution crawler started.")
    print(f"Script folder: {SCRIPT_DIR}")
    print(f"Output folder: {out_dir}")
    print(f"Output file will be: {out_path}")
    if update_analysis:
        print(f"Analysis workbook update: ON ({analysis_file})")
    else:
        print("Analysis workbook update: OFF")
    print("Collecting URLs...")
    link_rows = get_all_link_rows(headless=headless)

    matrix_rows = []
    req_rows = []
    upgrade_rows = []
    error_rows = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(viewport={"width": 1400, "height": 1200})

        for idx, link_row in enumerate(link_rows, start=1):
            print(f"[{idx}/{len(link_rows)}] {link_row['Source_Group']} | {link_row['URL']}")
            matrix, req_long, up_long, err = scrape_detail(context, link_row)
            if err:
                print("  ERROR:", err)
                error_rows.append({
                    "Source_Group": link_row.get("Source_Group"),
                    "Category": link_row.get("Category"),
                    "URL": link_row.get("URL"),
                    "Error": err,
                })
                continue
            if matrix is None:
                continue
            matrix_rows.append(matrix)
            req_rows.extend(req_long)
            upgrade_rows.extend(up_long)

        browser.close()

    df_matrix = pd.DataFrame(matrix_rows)
    df_req = pd.DataFrame(req_rows)
    df_up = pd.DataFrame(upgrade_rows)
    df_errors = pd.DataFrame(error_rows)
    df_links = pd.DataFrame(link_rows)

    success_count = len(matrix_rows)
    url_count = len(link_rows)
    error_count = len(error_rows)
    success_rate = (success_count / url_count) if url_count else 0
    print(f"Scrape result: success={success_count}, errors={error_count}, urls={url_count}, success_rate={success_rate:.1%}")

    severe_failure = url_count >= 10 and (success_count < MIN_SUCCESS_ROWS or success_rate < MIN_SUCCESS_RATE)
    if severe_failure:
        raise RuntimeError(
            "Severe scrape failure: too few detail pages parsed. "
            f"success={success_count}, urls={url_count}, errors={error_count}. "
            "Stop here to avoid overwriting latest analysis files with incomplete data."
        )

    first_cols = [
        "Source_Group", "Category", "Evolution_Name", "URL", "Cost_Coins", "Cost_FP",
        "Req_OVR_Max", "Req_OVR_Min", "Req_Position", "Req_Excluded_Position",
        "Req_Rarity", "Req_Excluded_Rarity",
        "Req_PAC_Max", "Req_SHO_Max", "Req_PAS_Max", "Req_DRI_Max", "Req_DEF_Max", "Req_PHY_Max",
        "Req_Max_PS", "Req_Max_PS_Plus", "Req_Max_Positions", "Req_Born_Before", "Req_Born_After",
    ]
    stat_cols = []
    for stat in ALL_STATS:
        stat_cols += [f"+ {stat}", f"Max {stat}"]
    tail_cols = ["Role_Added", "Other_Upgrade_Text", "Raw_Requirements_Text", "Req_Other", "Link_Text", "Parent_Text"]
    ordered = [c for c in first_cols + stat_cols + tail_cols if c in df_matrix.columns]
    rest = [c for c in df_matrix.columns if c not in ordered]
    if not df_matrix.empty:
        df_matrix = df_matrix[ordered + rest]

    # 기존 분석파일의 Unified_Matrix/Evo 표 자동 업데이트
    # 기본값은 OFF입니다. 로컬에서 필요한 경우 --update-analysis 옵션을 사용하세요.
    if update_analysis:
        if analysis_file.exists():
            try:
                update_analysis_workbook(df_matrix, analysis_file)
                print(f"분석파일 업데이트 완료: {analysis_file}")
            except PermissionError:
                print(f"분석파일 업데이트 실패: 파일이 열려 있습니다. Excel에서 닫고 다시 실행하세요. ({analysis_file})")
            except Exception as e:
                print(f"분석파일 업데이트 실패: {type(e).__name__}: {e}")
        else:
            print(f"분석파일을 찾지 못했습니다: {analysis_file}")
            print("--analysis 옵션으로 분석파일 경로를 지정하거나, 해당 파일을 출력 폴더에 넣어주세요.")

    summary_rows = []
    if not df_matrix.empty:
        for key, val in df_matrix["Source_Group"].value_counts(dropna=False).items():
            summary_rows.append({"Metric": f"{key}_Count", "Value": int(val)})

    summary_rows += [
        {"Metric": "Generated_At", "Value": datetime.now().isoformat(timespec="seconds")},
        {"Metric": "Start_URL", "Value": START_URL},
        {"Metric": "URL_Count", "Value": len(link_rows)},
        {"Metric": "Rule", "Value": "v8 portable/GitHub Actions-ready version; domcontentloaded navigation avoids networkidle timeouts."},
        {"Metric": "Rule", "Value": "Only true detail URLs like /evolutions/1962-rapid-gains/ are collected."},
        {"Metric": "Rule", "Value": "FACE stats such as PAC/SHO/PAS/DRI/DEF/PHY are not distributed into sub-stats."},
    ]
    df_summary = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_matrix.to_excel(writer, sheet_name="Unified_Matrix", index=False)
        df_up.to_excel(writer, sheet_name="Upgrades_Long", index=False)
        df_req.to_excel(writer, sheet_name="Requirements_Long", index=False)
        df_links.to_excel(writer, sheet_name="Collected_URLs", index=False)
        df_errors.to_excel(writer, sheet_name="Errors", index=False)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)

    print()
    print(f"Done: {out_path}")
    if not df_matrix.empty:
        print(df_matrix["Source_Group"].value_counts(dropna=False).to_string())


if __name__ == "__main__":
    main()
