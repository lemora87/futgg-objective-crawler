# futgg_objective_crawler_v1.py
# FUT.GG Objective crawler + mission condition parser + play plan generator
#
# Run:
#   py C:\Users\user\futgg_objective_crawler_v1.py --show
# or:
#   py C:\Users\user\futgg_objective_crawler_v1.py
#
# First-time setup, if needed:
#   py -m pip install pandas openpyxl beautifulsoup4 playwright
#   py -m playwright install chromium
#
# Output:
#   By default, output files are saved in the same folder as this .py file.
#   You can also choose another output folder with:
#     py futgg_objective_crawler_v20_portable.py --out "D:\\FUTGG"
#   or by creating a config file named futgg_objective_config.json next to this script.
#
# v32: add new-mission triggered combo reports that stack Any Ultimate + core SB/Rivals/Champions/Live Events + target mode.
#
# Purpose:
#   1) Active objective pages only
#   2) Mission_DB: mission-by-mission source table
#   3) Condition_DB: squad/performer/result/mode conditions, one condition per row
#   4) Mode_Plan: compact gameplay plan by mode, excluding Seasonal by default
#   5) Seasonal_Mode_Plan / Mode_Plan_All: optional reference views
#   6) Conflict_Check: basic conflict/attention flags
#
# Note:
#   FUT.GG objective text can change. This parser is rule-based and should be improved
#   whenever new wording appears. Raw_Description and Parse_Note columns are kept
#   so that incorrect parsing can be reviewed.

from __future__ import annotations

import re
import sys
import time
import json
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


START_URL = "https://www.fut.gg/objectives/"
BASE_URL = "https://www.fut.gg"
# Portable path settings
# - Default: save results to the same folder as this script.
# - Optional command-line override:
#     --out "D:\\FUTGG"
# - Optional config file next to this script:
#     futgg_objective_config.json
#     {
#       "out_dir": "D:/FUTGG",
#       "headless": false
#     }
SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_FILE = SCRIPT_DIR / "futgg_objective_config.json"
DEFAULT_OUT_DIR = SCRIPT_DIR
KST = timezone(timedelta(hours=9))
DEADLINE_BASE_HOUR_KST = 2
DEADLINE_BASE_MINUTE_KST = 0

# If you later make a permanent analysis workbook, set this to True and rename ANALYSIS_FILE.
UPDATE_ANALYSIS_FILE = False
ANALYSIS_FILE_NAME = "futgg_objectives_analysis.xlsx"

EXCLUDE_EXPIRED = True

# Seasonal은 장기 반복 미션이 많아서 기본 실행계획에서는 제외합니다.
# Mission_DB/Condition_DB에는 남기고, 별도 Seasonal_Mode_Plan 시트로 분리합니다.
EXCLUDE_SEASONAL_FROM_MAIN_PLAN = True

# Known names used only for attribute classification. Unknown values are still kept.
KNOWN_LEAGUES = {
    "Bundesliga", "Premier League", "LALIGA EA SPORTS", "LaLiga", "Serie A", "Ligue 1",
    "Eredivisie", "MLS", "ROSHN Saudi League", "EFL Championship", "Liga Portugal",
    "BWSL", "Barclays WSL", "Frauen-Bundesliga",
}
KNOWN_NATIONS = {
    "Germany", "German", "Spain", "Spanish", "Uruguay", "Scotland", "Korea Republic", "China PR",
    "USA", "United States", "England", "France", "Italy", "Brazil", "Argentina",
    "Netherlands", "Portugal", "Mexico", "Japan", "Nigeria", "Canada",
}
KNOWN_COMPETITIONS_OR_RARITIES = {
    "CONMEBOL Libertadores", "CONMEBOL Sudamericana", "TOTS", "TOTW", "Team of the Season",
    "World Tour Silver Stars", "Silver", "Gold", "Bronze",
}
POSITION_GROUPS = {"Defender", "Defenders", "Midfielder", "Midfielders", "Attacker", "Attackers"}
POSITIONS = {"GK", "LB", "LWB", "CB", "RB", "RWB", "CDM", "CM", "CAM", "LM", "RM", "LW", "RW", "ST", "CF"}


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def canonical_url(url: str) -> str:
    return (url or "").split("?", 1)[0].rstrip("/") + "/"


def is_objective_detail_url(url: str) -> bool:
    u = canonical_url(url)
    return re.search(r"/objectives/(campaigns|live-events|seasonal)/[0-9]+-[^/]+/$", u) is not None


def is_expired_text(text: str) -> bool:
    # Expires in ... is active; Expired is inactive.
    return re.search(r"\bExpired\b", clean_text(text), re.I) is not None


def parse_expires_text(text: str) -> str:
    """
    Extract active campaign deadline wording from FUT.GG card/detail text.
    Examples:
      'Expires in 6 days 5 Objectives ...' -> 'Expires in 6 days'
      'Expires in 17 hours 1,250 total ...' -> 'Expires in 17 hours'
      'Expires in 1 day 6 hours ...' -> 'Expires in 1 day 6 hours'
    """
    t = clean_text(text)
    if not t:
        return ""

    # Prefer duration units after "Expires in"; stop naturally before objective/progress text.
    m = re.search(
        r"\bExpires\s+in\s+((?:[0-9]+\s*(?:days?|hours?|hrs?|minutes?|mins?)\s*){1,3})",
        t,
        re.I,
    )
    if m:
        return "Expires in " + clean_text(m.group(1))

    # Fallback: capture a short phrase after Expires in.
    m = re.search(r"\bExpires\s+in\s+(.+?)(?=\s+[0-9,]+\s+total|\s+[0-9]+\s+Objectives|\s+[0-9]+%|$)", t, re.I)
    if m:
        return "Expires in " + clean_text(m.group(1))

    return ""


def scheduled_deadline_base_kst(now_dt: Optional[datetime] = None) -> datetime:
    """
    Use a fixed daily 02:00 KST base for 'Expires in ...' calculations.

    GitHub Actions scheduled jobs can start late. For example, a workflow scheduled
    for 02:00 KST might actually start later. If we used the actual start time,
    'Expires in 2 days' would shift incorrectly, which is misleading.

    Rule:
      - If current KST time is 02:00 or later: use today's 02:00 KST.
      - If current KST time is before 02:00: use yesterday's 02:00 KST.
    """
    if now_dt is None:
        now_dt = datetime.now(KST)
    elif now_dt.tzinfo is None:
        now_dt = now_dt.replace(tzinfo=KST)
    else:
        now_dt = now_dt.astimezone(KST)

    base = now_dt.replace(
        hour=DEADLINE_BASE_HOUR_KST,
        minute=DEADLINE_BASE_MINUTE_KST,
        second=0,
        microsecond=0,
    )

    if now_dt < base:
        base = base - timedelta(days=1)

    return base


def expiry_deadline_kst(expires_text: str, base_dt: Optional[datetime] = None) -> tuple[str, str]:
    """
    Convert relative FUT.GG wording into an absolute Korean-time deadline.
    Returns:
      (deadline_display, deadline_iso)

    Example:
      Expires in 2 days and scheduled base is 2026-05-05 02:00 KST
      -> ('5/7 02:00', '2026-05-07T02:00+09:00')

    Note:
      FUT.GG wording itself can be rounded, so this is an estimated deadline.
      The base time is fixed to daily 02:00 KST to avoid GitHub Actions delay noise.
    """
    text = clean_text(expires_text)
    if not text:
        return "", ""

    if base_dt is None:
        base_dt = scheduled_deadline_base_kst()
    elif base_dt.tzinfo is None:
        base_dt = base_dt.replace(tzinfo=KST)
    else:
        base_dt = base_dt.astimezone(KST)

    duration = timedelta(0)
    for num, unit in re.findall(r"([0-9]+)\s*(days?|hours?|hrs?|minutes?|mins?)", text, re.I):
        n = int(num)
        u = unit.lower()
        if u.startswith("day"):
            duration += timedelta(days=n)
        elif u.startswith("hour") or u.startswith("hr"):
            duration += timedelta(hours=n)
        elif u.startswith("minute") or u.startswith("min"):
            duration += timedelta(minutes=n)

    if duration == timedelta(0):
        return "", ""

    deadline = base_dt + duration
    return deadline.strftime("%-m/%-d %H:%M") if sys.platform != "win32" else deadline.strftime("%#m/%#d %H:%M"), deadline.isoformat(timespec="minutes")


def campaign_expiry_display(expires_text: str, deadline_kst: str) -> str:
    """
    Compact display used in report columns.
    Prefer absolute deadline; keep relative text as fallback.
    """
    deadline = clean_text(deadline_kst)
    if deadline:
        return f"마감 {deadline}"
    rel = clean_text(expires_text)
    return rel


def slug_title(url: str) -> str:
    slug = canonical_url(url).rstrip("/").split("/")[-1]
    slug = re.sub(r"^\d+-", "", slug)
    return " ".join(w.capitalize() for w in slug.split("-"))


def unique_by_url(rows: List[Dict]) -> List[Dict]:
    seen = set()
    out = []
    for r in rows:
        u = r.get("URL")
        if u and u not in seen:
            seen.add(u)
            out.append(r)
    return out


def collect_objective_links(headless: bool = True) -> List[Dict]:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        page = browser.new_page(viewport={"width": 1450, "height": 1200})
        page.goto(START_URL, wait_until="networkidle", timeout=60000)
        time.sleep(2)

        # Click expand buttons if any
        for _ in range(8):
            clicked = False
            for pat in [r"Show all", r"View all", r"Load more", r"Show more"]:
                try:
                    loc = page.get_by_text(re.compile(pat, re.I))
                    for i in range(min(loc.count(), 5)):
                        try:
                            t = loc.nth(i)
                            if t.is_visible(timeout=500):
                                t.click(timeout=1500, force=True)
                                clicked = True
                                time.sleep(0.4)
                        except Exception:
                            pass
                except Exception:
                    pass
            if not clicked:
                break
            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass

        raw = page.eval_on_selector_all(
            "a[href]",
            """
            els => els.map(a => {
                let parent = a;
                for (let i = 0; i < 6; i++) {
                    if (!parent.parentElement) break;
                    parent = parent.parentElement;
                    const txt = (parent.innerText || parent.textContent || '').trim();
                    if (txt.length > 40) break;
                }
                return {
                    href: a.href,
                    link_text: (a.innerText || a.textContent || '').trim(),
                    parent_text: (parent.innerText || parent.textContent || '').trim()
                };
            })
            """
        )

        rows = []
        for item in raw:
            url = canonical_url(item.get("href", ""))
            if not is_objective_detail_url(url):
                continue
            link_text = clean_text(item.get("link_text", ""))
            parent_text = clean_text(item.get("parent_text", ""))

            if EXCLUDE_EXPIRED and is_expired_text(parent_text + " " + link_text):
                continue

            campaign_expires = parse_expires_text(parent_text + " " + link_text)
            campaign_deadline_kst, campaign_deadline_iso = expiry_deadline_kst(campaign_expires)

            rows.append({
                "URL": url,
                "Link_Text": link_text,
                "Parent_Text": parent_text,
                "Objective_Type": infer_objective_type(url),
                "Campaign_Expires": campaign_expires,
                "Campaign_Deadline_KST": campaign_deadline_kst,
                "Campaign_Deadline_ISO": campaign_deadline_iso,
            })

        browser.close()

    rows = unique_by_url(rows)
    print(f"Collected active objective URLs: {len(rows)}")
    return rows


def infer_objective_type(url: str) -> str:
    if "/campaigns/" in url:
        return "Campaign"
    if "/live-events/" in url:
        return "Live Events"
    if "/seasonal/" in url:
        return "Seasonal"
    return "Objective"


def split_lines(text: str) -> List[str]:
    lines = []
    for line in (text or "").replace("\r", "\n").split("\n"):
        line = clean_text(line)
        if not line:
            continue
        # Remove obvious navigation noise
        if line.lower() in {"home", "objectives", "players", "sbc", "evolutions", "login", "sign up"}:
            continue
        lines.append(line)
    return lines


ACTION_START_RE = re.compile(
    r"^(Win|Play|Score|Assist|Complete|Keep|Earn|Make|Get|Perform|Record|Concede|Claim|Finish|Achieve)\b",
    re.I,
)


REWARD_LINE_RE = re.compile(
    r"\b(Pack|Player Pick|Player|SP|Coin Boost|Coins|Evo Unlock|XP|Token|Loan|EVO consumables?|Rare Gold|TOTS HM)\b",
    re.I,
)


def is_reward_line(line: str) -> bool:
    t = clean_text(line)
    if not t:
        return False
    # Short lines like "82+ Rare Gold Player Pack", "Coin Boost", "Evo Unlock"
    # are rewards, not objective mission names.
    return bool(REWARD_LINE_RE.search(t))


def split_description_and_reward(line: str) -> tuple[str, str]:
    """
    FUT.GG sometimes combines the next mission description and reward in one text line:
      Keep 4 clean sheets ... starting 11. / 82+ Rare Gold Player Pack / Evo Unlock

    For parsing, keep only the mission sentence as Raw_Description and move the
    suffix after "/" into Reward.
    """
    t = clean_text(line)
    if " / " not in t:
        return t, ""

    parts = [clean_text(p) for p in t.split(" / ") if clean_text(p)]
    if not parts:
        return t, ""

    # If the first part is an action sentence, treat remaining reward-looking parts as reward.
    if ACTION_START_RE.search(parts[0]):
        rewards = [p for p in parts[1:] if is_reward_line(p)]
        return parts[0], " / ".join(rewards)

    return t, ""


def reward_from_nearby_lines(lines: List[str], start_idx: int) -> str:
    """
    Reward lines near a mission description. Do not treat full mission
    descriptions as rewards even if they contain reward words after slash.
    """
    found = []
    for j in range(start_idx + 1, min(start_idx + 7, len(lines))):
        cand = clean_text(lines[j])
        desc_part, reward_part = split_description_and_reward(cand)

        if reward_part and desc_part != cand:
            # Candidate is actually another mission description with reward suffix.
            # Do not attach it to the current mission.
            continue

        if is_reward_line(cand) and not looks_like_description(cand) and not looks_like_mission_name(cand):
            found.append(cand)

    return " / ".join(found[:3])


def looks_like_description(line: str) -> bool:
    """
    A mission description should usually be a full sentence, not just a card title
    like 'Assist 4' or a pure reward line like '82+ Rare Gold Player Pack'.

    If a line is "description / reward", judge only the description part.
    """
    t, _ = split_description_and_reward(line)
    t = clean_text(t)
    if not t:
        return False

    # Pure reward lines are not descriptions. But a description with reward suffix
    # was already reduced to the first action sentence above.
    if is_reward_line(t) and not ACTION_START_RE.search(t):
        return False

    # Very short action labels are mission names, not descriptions.
    if re.fullmatch(r"(Win|Play|Score|Assist|Keep|Clean Sheet)\s+(?:in\s+)?[0-9]+", t, re.I):
        return False

    has_action = bool(ACTION_START_RE.search(t))
    has_context = bool(re.search(
        r"\b(matches?|goals?|assists?|clean sheets?|Squad Battles|Rivals|Champions|Live Events?|Rush|starting 11|using|with|while having|difficulty)\b",
        t,
        re.I,
    ))
    return has_action and has_context


def looks_like_mission_name(line: str) -> bool:
    t = clean_text(line)
    if len(t) > 60:
        return False
    if re.search(r"https?://", t):
        return False
    if t in {"Rewards", "Requirements", "Objectives", "Squad", "Details"}:
        return False
    if is_reward_line(t):
        return False

    # Names often look like "Win 2", "Score in 2", "Play 5", "Assist 4".
    if re.search(r"^(Win|Play|Score|Assist|Keep|Clean Sheet|Complete)\b", t, re.I):
        return True

    # Some cards have short custom names.
    return bool(re.search(r"[A-Za-z]", t)) and not looks_like_description(t)


def extract_reward_near(lines: List[str], start_idx: int) -> str:
    # Backward-compatible wrapper, but safer than the older implementation.
    return reward_from_nearby_lines(lines, start_idx)


def extract_missions_from_text(text: str) -> List[Dict]:
    """
    Extract mission cards robustly.

    FUT.GG objective pages often appear as:
      Mission title
      Reward
      Full mission description

    The older pairwise parser sometimes treated reward lines as mission names
    and short titles like 'Assist 4' as descriptions. This version finds full
    description lines first, then looks backward for the nearest mission title.
    """
    lines = split_lines(text)
    missions = []
    seen_desc = set()

    for i, line in enumerate(lines):
        raw_line = clean_text(line)
        desc, reward_suffix = split_description_and_reward(raw_line)
        desc = clean_text(desc)

        if not looks_like_description(raw_line):
            continue
        if desc in seen_desc:
            continue
        seen_desc.add(desc)

        # Look backward for the nearest valid short mission title.
        name = ""
        for j in range(i - 1, max(-1, i - 5), -1):
            cand = clean_text(lines[j])
            if looks_like_mission_name(cand):
                name = cand
                break

        if not name:
            # Fallback to action + count from the description.
            action, count, _, _ = parse_action(desc, "")
            name = f"{action} {count}" if action != "Other" else desc[:40]

        nearby_reward = extract_reward_near(lines, i)
        reward = " / ".join([x for x in [reward_suffix, nearby_reward] if x])

        missions.append({
            "Mission_Name": name,
            "Raw_Description": desc,
            "Reward": reward,
            "Extract_Method": "description_line_with_backward_title",
        })

    # Fallback: if no full descriptions detected, use action lines.
    if not missions:
        seen = set()
        for i, line in enumerate(lines):
            if ACTION_START_RE.search(line) and not is_reward_line(line):
                key = clean_text(line)
                if key in seen:
                    continue
                seen.add(key)
                missions.append({
                    "Mission_Name": key[:40],
                    "Raw_Description": key,
                    "Reward": extract_reward_near(lines, i),
                    "Extract_Method": "action_line_fallback",
                })

    return missions


def classify_value(value: str) -> str:
    v = clean_text(value)
    if v in POSITIONS:
        return "Position"
    if v in POSITION_GROUPS:
        return "Position_Group"
    if v in KNOWN_LEAGUES:
        return "League"
    if v in KNOWN_NATIONS:
        return "Nation"
    if v in KNOWN_COMPETITIONS_OR_RARITIES:
        return "Rarity/Competition"
    # Common suffix clue
    if re.search(r"League|Liga|Bundesliga|LALIGA|Serie|Ligue", v, re.I):
        return "League"
    if re.search(r"Libertadores|Sudamericana|TOTS|TOTW|Silver|Gold|Bronze", v, re.I):
        return "Rarity/Competition"
    return "Unknown"


def normalize_value(value: str) -> str:
    v = clean_text(value)
    aliases = {
        "German": "Germany",
        "Spanish": "Spain",
        "American": "USA",
        "Korean": "Korea Republic",
        "Chinese": "China PR",
        "Defenders": "Defender",
        "Midfielders": "Midfielder",
        "Attackers": "Attacker",
    }
    return aliases.get(v, v)



def normalize_mode_group(mode_group: str) -> str:
    """
    FUT.GG sometimes writes similar eligible-mode groups slightly differently:
      Squad Battles / Rivals / Champions / Live Events
      Squad Battles / Rivals / Champions / Live Events / Rush

    For practical play-planning, merge these into one bucket.
    Rush availability can still be checked in Mission_DB Raw_Description if needed.
    """
    mg = clean_text(mode_group)

    parts = {p.strip().lower() for p in mg.split("/") if p.strip()}
    core = {"squad battles", "rivals", "champions", "live events"}
    if core.issubset(parts):
        return "Squad Battles / Rivals / Champions / Live Events"

    return mg


def parse_mode_and_difficulty(desc: str, campaign_title: str = "") -> Tuple[str, str]:
    d = clean_text(desc)
    modes = []
    difficulty = ""

    if re.search(r"Squad Battles", d, re.I):
        modes.append("Squad Battles")
    if re.search(r"Rivals", d, re.I):
        modes.append("Rivals")
    if re.search(r"Champions", d, re.I):
        modes.append("Champions")
    if re.search(r"Live Events?", d, re.I):
        modes.append("Live Events")
    if re.search(r"\bRush\b", d, re.I):
        modes.append("Rush")

    if re.search(r"any Ultimate Team", d, re.I):
        modes.append("Any Ultimate Team mode")

    m = re.search(r"Min\.?\s*([A-Za-z -]+?)\s+difficulty", d, re.I)
    if m:
        difficulty = "Min. " + clean_text(m.group(1))

    # Event-specific pages often say "in the [campaign] event"
    ev = re.search(r"in the (.+? event)\b", d, re.I)
    if ev:
        modes = [clean_text(ev.group(1))]

    if not modes and "Rush" in campaign_title:
        modes = [campaign_title]
    if not modes and "Exhibition" in campaign_title:
        modes = [campaign_title]
    if not modes:
        modes = ["Unspecified / Check Raw"]

    return " / ".join(unique_list(modes)), difficulty


def unique_list(items: List[str]) -> List[str]:
    seen = set()
    out = []
    for x in items:
        x = clean_text(x)
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def is_goal_target(target: str) -> bool:
    """
    Report classification helper.
    Treat special goal wording as Goals, not Other_Actions.
    Examples: Outside Box Goal, Header Goal, Low Driven Goal, Hat Trick.
    """
    t = clean_text(target).lower()
    if not t:
        return False
    return (
        "goal" in t
        or "score" in t
        or "hat trick" in t
        or "hat-trick" in t
    )


def is_assist_target(target: str) -> bool:
    """
    Report classification helper.
    Treat special assist wording as Assists, not Other_Actions.
    Examples: Through Ball Assist, Cross Assist, Incisive Pass Assist.
    """
    t = clean_text(target).lower()
    if not t:
        return False
    return "assist" in t


def detect_special_action(desc: str, base_action: str) -> str:
    """
    Detect special goal/assist/action wording that should be surfaced in Other_Actions.
    Examples:
      - from outside the box -> Outside Box Goal
      - Finesse goals -> Finesse Goal
      - Low Driven -> Low Driven Goal
      - Through Ball assists -> Through Ball Assist
      - headers -> Header Goal
      - volleys -> Volley Goal
    """
    d = clean_text(desc)

    # Goal subtypes
    if base_action == "Goal":
        if re.search(r"outside\s+the\s+box|from\s+distance|long[- ]range", d, re.I):
            return "Outside Box Goal"
        if re.search(r"\bfinesse\b", d, re.I):
            return "Finesse Goal"
        if re.search(r"low\s+driven", d, re.I):
            return "Low Driven Goal"
        if re.search(r"power\s+shot", d, re.I):
            return "Power Shot Goal"
        if re.search(r"chip\s+shot", d, re.I):
            return "Chip Shot Goal"
        if re.search(r"\bheader(s)?\b|headed\s+goals?", d, re.I):
            return "Header Goal"
        if re.search(r"\bvolley(s)?\b", d, re.I):
            return "Volley Goal"
        if re.search(r"hat[- ]?trick", d, re.I):
            return "Hat Trick"

    # Assist subtypes
    if base_action == "Assist":
        if re.search(r"through\s+ball", d, re.I):
            return "Through Ball Assist"
        if re.search(r"lobbed\s+through", d, re.I):
            return "Lobbed Through Ball Assist"
        if re.search(r"incisive\s+pass", d, re.I):
            return "Incisive Pass Assist"
        if re.search(r"inventive\s+pass", d, re.I):
            return "Inventive Pass Assist"
        if re.search(r"flair\s+pass", d, re.I):
            return "Flair Pass Assist"
        if re.search(r"cross(es)?\b|from\s+crosses", d, re.I):
            return "Cross Assist"

    # Defensive/special actions
    if re.search(r"\btackles?\b", d, re.I):
        return "Tackle"
    if re.search(r"\binterceptions?\b", d, re.I):
        return "Interception"
    if re.search(r"concede\s+no\s+more\s+than|concede\s+(?:less|fewer)\s+than", d, re.I):
        return "Concede Limit"
    if re.search(r"score\s+and\s+assist|goal\s+and\s+assist|goals?\s*&\s*assists?", d, re.I):
        return "Goal+Assist Same Match"

    return base_action



def parse_action(desc: str, mission_name: str = "") -> Tuple[str, int, str, str]:
    d = clean_text(desc)
    n = clean_text(mission_name)

    # FUT objective wording에서 separate가 나오면 사실상 경기별 조건으로 봅니다.
    per_match = "Yes" if re.search(r"\bseparate\b|each\s+match|per\s+match", d, re.I) else "No"
    requires_win = "Yes" if re.search(r"^Win\b|\bWin\s+\d+|\bwhile winning\b", d, re.I) or re.search(r"^Win\b", n, re.I) else "No"

    # Important: separate-match wording should use the number of matches,
    # not the number of goals in each match.
    # Examples:
    #   Score 1 goal in 5 separate matches -> count 5, per_match Yes
    #   Score at least 1 goal in 2 separate matches -> count 2, per_match Yes
    #   Assist at least 1 goal in 5 separate matches -> count 5, per_match Yes
    separate_patterns = [
        # "separate Squad Battles matches"처럼 separate와 matches 사이에 모드명이 들어가는 경우까지 허용
        ("Assist", r"\bAssist(?:\s+at\s+least)?\s+(?:1|one)\s+goals?\s+in\s+([0-9]+)\s+separate\s+(?:\w+\s+){0,4}matches?\b"),
        ("Assist", r"\bAssist\s+in\s+([0-9]+)\s+separate\s+(?:\w+\s+){0,4}matches?\b"),
        ("Goal", r"\bScore(?:\s+at\s+least)?\s+(?:1|one|a)\s+goals?\s+in\s+([0-9]+)\s+separate\s+(?:\w+\s+){0,4}matches?\b"),
        ("Goal", r"\bScore\s+(?:in\s+)?([0-9]+)\s+separate\s+(?:\w+\s+){0,4}matches?\b"),
    ]

    for action, pat in separate_patterns:
        m = re.search(pat, d, re.I)
        if m:
            action = detect_special_action(d, action)
            return action, int(m.group(1)), "Yes", requires_win

    patterns = [
        ("Win", r"\bWin\s+([0-9]+)\b"),
        ("Play", r"\bPlay\s+([0-9]+)\b"),
        ("Assist", r"\bAssist(?:\s+at least\s+1\s+goal\s+in)?\s*([0-9]+)\b"),
        ("Assist", r"\b(?:Make|Record|Get|Perform)\s+([0-9]+)\s+assists?\b"),
        ("Goal", r"\bScore(?:\s+at least\s+1\s+goal\s+in)?\s*([0-9]+)\b"),
        ("Goal", r"\b(?:Get|Make|Record)\s+([0-9]+)\s+goals?\b"),
        ("Clean Sheet", r"\b(?:Keep|Record)\s+([0-9]+)\s+clean sheets?\b"),
        ("Clean Sheet", r"\b([0-9]+)\s+clean sheets?\b"),
        ("Complete", r"\bComplete\s+([0-9]+)\b"),
    ]

    for action, pat in patterns:
        m = re.search(pat, d, re.I)
        if m:
            action = detect_special_action(d, action)
            return action, int(m.group(1)), per_match, requires_win

    # Mission name fallback
    for action, pat in patterns:
        m = re.search(pat, n, re.I)
        if m:
            action = detect_special_action(d + " " + n, action)
            return action, int(m.group(1)), per_match, requires_win

    return detect_special_action(d + " " + n, "Other"), 1, per_match, requires_win


def strip_trailing_context(value: str) -> str:
    v = clean_text(value)
    # Stop at common tail words
    v = re.split(
        r"\b(in your starting 11|in the starting 11|while|on Min|difficulty|in Squad Battles|in Rivals|in Champions|in Live Events|in Rush|using|with)\b",
        v,
        flags=re.I,
    )[0]
    v = re.sub(r"\.$", "", clean_text(v))
    return v


def parse_squad_conditions(desc: str, mission_id: str) -> List[Dict]:
    d = clean_text(desc)
    rows = []

    # Examples:
    # while having Min. 1 player from the Bundesliga in your starting 11
    # while having Min. 2 Uruguay players in your starting 11
    patterns = [
        r"(?:while\s+)?having\s+min\.?\s*([0-9]+)\s+players?\s+from\s+(?:the\s+)?(.+?)\s+in\s+your\s+starting\s+11",
        r"(?:while\s+)?having\s+min\.?\s*([0-9]+)\s+(.+?)\s+players?\s+in\s+your\s+starting\s+11",
        r"min\.?\s*([0-9]+)\s+players?\s+from\s+(?:the\s+)?(.+?)\s+in\s+your\s+starting\s+11",
        r"min\.?\s*([0-9]+)\s+(.+?)\s+players?\s+in\s+your\s+starting\s+11",
        # FUT.GG sometimes says: "while having min. 1 CONMEBOL Libertadores player in your starting 11"
        r"(?:while\s+)?having\s+min\.?\s*([0-9]+)\s+(.+?)\s+player\s+in\s+your\s+starting\s+11",
        r"min\.?\s*([0-9]+)\s+(.+?)\s+player\s+in\s+your\s+starting\s+11",
    ]

    for pat in patterns:
        for m in re.finditer(pat, d, re.I):
            count = int(m.group(1))
            value = strip_trailing_context(m.group(2))
            value = normalize_value(value)
            if not value or len(value) > 70:
                continue
            rows.append({
                "Mission_ID": mission_id,
                "Condition_Type": "Squad",
                "Target": "Starting XI",
                "Attribute": classify_value(value),
                "Value": value,
                "Count": count,
                "Position": "",
                "Per_Match": "No",
                "Operational_Note": "선발에 포함",
            })

    # Position-specific rough pattern: "as LB" / "at LB"
    # This is conservative and may need updating with real examples.
    pos = re.search(r"\b(?:as|at)\s+(GK|LB|LWB|CB|RB|RWB|CDM|CM|CAM|LM|RM|LW|RW|ST|CF)\b", d, re.I)
    if pos and rows:
        rows[-1]["Position"] = pos.group(1).upper()

    # Remove duplicates
    return dedupe_condition_rows(rows)


def parse_performer_conditions(desc: str, mission_id: str, action: str, count: int, per_match: str) -> List[Dict]:
    d = clean_text(desc)
    rows = []

    # Examples:
    # using a German player
    # using any TOTS player
    # using a Spanish player
    # with players from China PR
    # with Defender
    patterns = [
        # using a Defender (Preferred Position)
        r"\busing\s+(?:a|an)\s+([A-Za-z ]+?)\s*\(Preferred Position\)",
        # using a Defender
        r"\busing\s+(?:a|an)\s+(Defender|Midfielder|Attacker|GK|LB|LWB|CB|RB|RWB|CDM|CM|CAM|LM|RM|LW|RW|ST|CF)\b",
        # using a player from Scotland / the BWSL / Frauen-Bundesliga / Barclays WSL / CONMEBOL Libertadores
        r"\busing\s+(?:a|an|any)\s+players?\s+from\s+(?:the\s+)?(.+?)(?:\.|$)",
        r"\busing\s+players?\s+from\s+(?:the\s+)?(.+?)(?:\.|$)",
        r"\bwith\s+(?:a|an|any)\s+players?\s+from\s+(?:the\s+)?(.+?)(?:\.|$)",
        r"\bwith\s+players?\s+from\s+(?:the\s+)?(.+?)(?:\.|$)",
        # using any TOTS player / using German players
        r"\busing\s+(?:any\s+)?(?:a|an)?\s*([A-Za-z0-9 +.'-]+?)\s+players?\b",
        r"\bwith\s+(?:a|an)?\s*([A-Za-z0-9 +.'-]+?)\s+players?\b",
    ]

    found = []

    # Explicit high-confidence catches first.
    explicit_from = [
        ("Scotland", r"\busing\s+(?:a|an|any)\s+players?\s+from\s+(?:the\s+)?Scotland\b"),
        ("BWSL", r"\busing\s+(?:a|an|any)\s+players?\s+from\s+(?:the\s+)?BWSL\b"),
        ("Frauen-Bundesliga", r"\busing\s+(?:a|an|any)\s+players?\s+from\s+(?:the\s+)?Frauen-Bundesliga\b"),
        ("Barclays WSL", r"\busing\s+(?:a|an|any)\s+players?\s+from\s+(?:the\s+)?Barclays\s+WSL\b"),
        ("CONMEBOL Libertadores", r"\busing\s+(?:a|an|any)\s+players?\s+from\s+(?:the\s+)?CONMEBOL\s+Libertadores\b"),
    ]
    for val, pat in explicit_from:
        if re.search(pat, d, re.I):
            found.append(val)

    if re.search(r"\busing\s+players?\s+from\s+(?:the\s+)?CONMEBOL\s+Libertadores\b", d, re.I):
        found.append("CONMEBOL Libertadores")
    if re.search(r"\bwith\s+players?\s+from\s+(?:the\s+)?CONMEBOL\s+Libertadores\b", d, re.I):
        found.append("CONMEBOL Libertadores")
    if re.search(r"\busing\s+players?\s+from\s+(?:the\s+)?CONMEBOL\s+Sudamericana\b", d, re.I):
        found.append("CONMEBOL Sudamericana")
    if re.search(r"\bwith\s+players?\s+from\s+(?:the\s+)?CONMEBOL\s+Sudamericana\b", d, re.I):
        found.append("CONMEBOL Sudamericana")

    for pat in patterns:
        for m in re.finditer(pat, d, re.I):
            val = strip_trailing_context(m.group(1))
            val = re.sub(r"\(Preferred Position\)", "", val, flags=re.I)
            val = re.sub(r"\b(any|a|an)\b", "", val, flags=re.I)
            val = normalize_value(clean_text(val))
            if val and len(val) <= 70:
                found.append(val)

    # Special "using any TOTS player" may be captured as TOTS.
    if re.search(r"using\s+any\s+TOTS\s+player", d, re.I):
        found.append("TOTS")

    # If no performer is specified but action is goal/assist or a special subtype, performer Any.
    if not found and (
        action in {"Goal", "Assist", "Outside Box Goal", "Finesse Goal", "Low Driven Goal", "Header Goal", "Volley Goal",
                   "Through Ball Assist", "Cross Assist", "Tackle", "Interception", "Goal+Assist Same Match",
                   "Power Shot Goal", "Chip Shot Goal", "Hat Trick", "Perfect Hat Trick"}
        or "Goal" in action
        or "Assist" in action
    ):
        found = ["Any"]

    for val in unique_list(found):
        rows.append({
            "Mission_ID": mission_id,
            "Condition_Type": "Performer",
            "Target": action,
            "Attribute": classify_value(val) if val != "Any" else "Any",
            "Value": val,
            "Count": count,
            "Position": "",
            "Per_Match": per_match,
            "Operational_Note": performer_note(action, count, per_match, val),
        })

    return dedupe_condition_rows(rows)


def performer_note(action: str, count: int, per_match: str, val: str) -> str:
    if per_match == "Yes":
        return f"{val}: {count}경기 각각 {action} 1회 이상"
    return f"{val}: {action} {count}회"


def parse_result_conditions(desc: str, mission_id: str, action: str, count: int, per_match: str, requires_win: str) -> List[Dict]:
    rows = []
    if action in {"Win", "Play", "Clean Sheet", "Complete"}:
        rows.append({
            "Mission_ID": mission_id,
            "Condition_Type": "Match_Result" if action in {"Win", "Clean Sheet"} else "Match_Count",
            "Target": action,
            "Attribute": "Any",
            "Value": "Any",
            "Count": count,
            "Position": "",
            "Per_Match": per_match,
            "Operational_Note": f"{action} {count}",
        })
    elif requires_win == "Yes":
        rows.append({
            "Mission_ID": mission_id,
            "Condition_Type": "Match_Result",
            "Target": "Win",
            "Attribute": "Any",
            "Value": "Any",
            "Count": count,
            "Position": "",
            "Per_Match": per_match,
            "Operational_Note": f"Win required",
        })
    return rows


def dedupe_condition_rows(rows: List[Dict]) -> List[Dict]:
    seen = set()
    out = []
    for r in rows:
        key = (r.get("Mission_ID"), r.get("Condition_Type"), r.get("Target"), r.get("Attribute"), r.get("Value"), r.get("Count"), r.get("Position"))
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out


def should_show_raw_in_other_actions(desc: str, conditions: List[Dict]) -> bool:
    """
    If a mission contains special-action-like wording but only generic parsing occurred,
    show the raw mission sentence in Mode_Plan.Other_Actions so it is not missed.
    """
    raw = clean_text(desc)
    if not raw:
        return False

    suspicious_patterns = [
        r"outside\s+the\s+box", r"from\s+distance", r"long[- ]range",
        r"\bfinesse\b", r"low\s+driven", r"through\s+ball",
        r"\bheader(s)?\b", r"headed\s+goals?", r"\bvolley(s)?\b",
        r"\btackles?\b", r"\binterceptions?\b", r"concede\s+no\s+more\s+than",
        r"concede\s+(?:less|fewer)\s+than", r"score\s+and\s+assist",
        r"goal\s+and\s+assist", r"goals?\s*&\s*assists?",
        r"hat[- ]?trick", r"weak\s+foot", r"skill\s+moves?",
        r"power\s+shot", r"chip\s+shot", r"trivela", r"precision\s+pass",
        r"lobbed\s+through", r"flair", r"first\s+time", r"volley",
    ]

    if not any(re.search(p, raw, re.I) for p in suspicious_patterns):
        return False

    # If a non-generic special target was already parsed, no raw fallback needed.
    generic_targets = {"Goal", "Assist", "Win", "Play", "Clean Sheet", "Complete", "Other"}
    parsed_targets = {str(c.get("Target", "")) for c in conditions}
    has_special = any(t and t not in generic_targets for t in parsed_targets)
    return not has_special



def parse_mission_conditions(mission: Dict) -> Tuple[Dict, List[Dict]]:
    desc = mission["Raw_Description"]
    title = mission.get("Campaign", "")
    mission_name = mission.get("Mission_Name", "")

    mode_group, difficulty = parse_mode_and_difficulty(desc, title)
    mode_group = normalize_mode_group(mode_group)
    action, count, per_match, requires_win = parse_action(desc, mission_name)

    mission_row = {
        "Mission_ID": mission["Mission_ID"],
        "Campaign": mission["Campaign"],
        "Campaign_Expires": mission.get("Campaign_Expires", ""),
        "Campaign_Deadline_KST": mission.get("Campaign_Deadline_KST", ""),
        "Campaign_Deadline_ISO": mission.get("Campaign_Deadline_ISO", ""),
        "Objective_Type": mission["Objective_Type"],
        "Mission_Name": mission["Mission_Name"],
        "Raw_Description": desc,
        "Mode_Group": mode_group,
        "Difficulty": difficulty,
        "Action": action,
        "Action_Count": count,
        "Per_Match": per_match,
        "Requires_Win": requires_win,
        "Reward": mission.get("Reward", ""),
        "Source_URL": mission["URL"],
        "Parse_Method": mission.get("Extract_Method", ""),
        "Parse_Note": "",
    }

    conditions = []
    conditions.extend(parse_squad_conditions(desc, mission["Mission_ID"]))
    conditions.extend(parse_performer_conditions(desc, mission["Mission_ID"], action, count, per_match))
    conditions.extend(parse_result_conditions(desc, mission["Mission_ID"], action, count, per_match, requires_win))

    # Safety rules for common FUT.GG wording that regexes may miss.
    # Example: "Score 5 goals ... using players from CONMEBOL Libertadores"
    if re.search(r"\busing\s+players?\s+from\s+(?:the\s+)?CONMEBOL\s+Libertadores\b", desc, re.I):
        exists = any(c["Condition_Type"] == "Performer" and c["Value"] == "CONMEBOL Libertadores" for c in conditions)
        if not exists and action in {"Goal", "Assist"}:
            conditions.append({
                "Mission_ID": mission["Mission_ID"],
                "Condition_Type": "Performer",
                "Target": action,
                "Attribute": "Rarity/Competition",
                "Value": "CONMEBOL Libertadores",
                "Count": count,
                "Position": "",
                "Per_Match": per_match,
                "Operational_Note": performer_note(action, count, per_match, "CONMEBOL Libertadores"),
            })

    # Example: "while having min. 1 CONMEBOL Libertadores player in your starting 11"
    m_con_squad = re.search(r"having\s+min\.?\s*([0-9]+)\s+CONMEBOL\s+Libertadores\s+players?\s+in\s+your\s+starting\s+11", desc, re.I)
    if m_con_squad:
        exists = any(c["Condition_Type"] == "Squad" and c["Value"] == "CONMEBOL Libertadores" for c in conditions)
        if not exists:
            conditions.append({
                "Mission_ID": mission["Mission_ID"],
                "Condition_Type": "Squad",
                "Target": "Starting XI",
                "Attribute": "Rarity/Competition",
                "Value": "CONMEBOL Libertadores",
                "Count": int(m_con_squad.group(1)),
                "Position": "",
                "Per_Match": "No",
                "Operational_Note": "선발에 포함",
            })

    conditions = dedupe_condition_rows(conditions)

    # Safety: if raw text contains outside-the-box wording, make sure it is visible in Other_Actions.
    if re.search(r"outside\s+the\s+box|from\s+distance|long[- ]range", desc, re.I):
        if not any(c["Condition_Type"] == "Performer" and c["Target"] == "Outside Box Goal" for c in conditions):
            # Reuse the best performer value if available, otherwise Any.
            performer_vals = [c["Value"] for c in conditions if c["Condition_Type"] == "Performer" and c["Target"] in {"Goal", "Outside Box Goal"}]
            val = performer_vals[0] if performer_vals else "Any"
            conditions.append({
                "Mission_ID": mission["Mission_ID"],
                "Condition_Type": "Performer",
                "Target": "Outside Box Goal",
                "Attribute": classify_value(val) if val != "Any" else "Any",
                "Value": val,
                "Count": count,
                "Position": "",
                "Per_Match": per_match,
                "Operational_Note": performer_note("Outside Box Goal", count, per_match, val),
            })

    # Final safety net: if special wording was not confidently parsed,
    # put the raw sentence into Other_Actions instead of hiding it in Parse_Check only.
    if should_show_raw_in_other_actions(desc, conditions):
        conditions.append({
            "Mission_ID": mission["Mission_ID"],
            "Condition_Type": "Other_Action",
            "Target": "Raw",
            "Attribute": "Raw",
            "Value": "Raw",
            "Count": count,
            "Position": "",
            "Per_Match": per_match,
            "Operational_Note": "원문 확인: " + desc,
        })

    conditions = dedupe_condition_rows(conditions)

    if not conditions:
        mission_row["Parse_Note"] = "No structured condition parsed; check Raw_Description."
    elif any(c["Attribute"] == "Unknown" for c in conditions):
        mission_row["Parse_Note"] = "Some condition values classified as Unknown; check manually."

    return mission_row, conditions


def scrape_objective_page(context, link_row: Dict, seq: int) -> Tuple[List[Dict], str]:
    url = link_row["URL"]
    page = context.new_page()
    try:
        page.goto(url, wait_until="networkidle", timeout=60000)
        time.sleep(1)

        body = page.inner_text("body", timeout=12000)
        if EXCLUDE_EXPIRED and is_expired_text(body):
            print(f"  SKIP expired page: {url}")
            return [], ""

        title = ""
        try:
            h1s = page.locator("h1").all_inner_texts()
            if h1s:
                title = clean_text(h1s[0])
        except Exception:
            pass
        if not title:
            title = slug_title(url)

        campaign_expires = link_row.get("Campaign_Expires", "") or parse_expires_text(body)
        campaign_deadline_kst = link_row.get("Campaign_Deadline_KST", "")
        campaign_deadline_iso = link_row.get("Campaign_Deadline_ISO", "")
        if not campaign_deadline_kst and campaign_expires:
            campaign_deadline_kst, campaign_deadline_iso = expiry_deadline_kst(campaign_expires)

        extracted = extract_missions_from_text(body)
        mission_rows = []
        for i, m in enumerate(extracted, start=1):
            mission_rows.append({
                "Mission_ID": f"O{seq:03d}-{i:02d}",
                "Campaign": title,
                "Campaign_Expires": campaign_expires,
                "Campaign_Deadline_KST": campaign_deadline_kst,
                "Campaign_Deadline_ISO": campaign_deadline_iso,
                "Objective_Type": link_row.get("Objective_Type", ""),
                "Mission_Name": m["Mission_Name"],
                "Raw_Description": m["Raw_Description"],
                "Reward": m.get("Reward", ""),
                "URL": url,
                "Extract_Method": m.get("Extract_Method", ""),
            })

        return mission_rows, clean_text(body)[:5000]
    except Exception as e:
        print(f"  ERROR detail: {type(e).__name__}: {e}")
        return [], ""
    finally:
        page.close()



def korean_value_label(value: str) -> str:
    """
    Korean display labels for report-facing result objectives.
    This intentionally covers common FUT.GG objective values. Unknown values are kept as-is.
    """
    v = clean_text(str(value or ""))
    aliases = {
        "Germany": "독일",
        "German": "독일",
        "USA": "미국",
        "United States": "미국",
        "Saudi Arabia": "사우디아라비아",
        "Saudi Pro League": "사우디 프로리그",
        "ROSHN Saudi Pro League": "사우디 프로리그",
        "ROSHN Saudi League": "사우디 리그",
        "Bundesliga": "분데스리가",
        "Frauen-Bundesliga": "프라우엔-분데스리가",
        "Ligue 1": "리그앙",
        "NWSL": "NWSL",
        "TOTS": "TOTS 카드",
        "French": "프랑스",
        "France": "프랑스",
        "Midfielder": "미드필더",
        "Defender": "수비수",
        "Attacker": "공격수",
        "CONMEBOL Libertadores": "코파 리베르타도레스",
        "Uruguay": "우루과이",
        "Swedish": "스웨덴",
        "Scotland": "스코틀랜드",
        "Barclays WSL": "바클레이스 WSL",
        "BWSL": "BWSL",
        "Any": "아무 선수",
        "-": "",
    }
    return aliases.get(v, v)


def korean_squad_requirement_line(row: Dict) -> str:
    """
    Convert a Squad condition row into a short Korean phrase:
      USA ≥2 -> 미국 선수 2명 이상 선발
      TOTS ≥1 -> TOTS 카드 1명 이상 선발
    """
    value = korean_value_label(row.get("Value", ""))
    count = int(row.get("Count") or 0)
    pos = clean_text(str(row.get("Position", "")))

    if not value:
        return ""

    if pos:
        return f"{pos} 포지션 {value} 선수 {count}명 이상 선발"

    # TOTS 카드/리그/국가 모두 자연스럽게 보이도록 '선수'를 붙임.
    if value.endswith("카드"):
        return f"{value} {count}명 이상 선발"
    return f"{value} 선수 {count}명 이상 선발"


def result_action_korean(target: str, count: int, per_match: str = "No") -> str:
    """
    Korean display for result/count objectives.
    """
    t = clean_text(str(target or ""))
    c = int(count or 0)

    if t == "Win":
        return f"{c}승"
    if t == "Play":
        return f"{c}경기 플레이"
    if t == "Complete":
        return f"{c}회 완료"
    if t == "Clean Sheet":
        return f"클린시트 {c}회"

    return f"{t} {c}"


def mission_result_todo_lines(cond_rows: List[Dict]) -> List[str]:
    """
    Every mission-level result/count objective should appear in To do(기타),
    not only in the top '플레이 요건' summary.

    Examples:
      Squad: USA ≥2 + Match_Result Win 7
        -> 미국 선수 2명 이상 선발하고 7승
      Squad: Ligue 1 ≥1 + Match_Count Play 2
        -> 리그앙 선수 1명 이상 선발하고 2경기 플레이
      No squad + Match_Count Play 10
        -> 10경기 플레이

    This preserves the per-objective connection between Starting XI conditions
    and Win/Play/Complete requirements.
    """
    squad_rows = [r for r in cond_rows if r.get("Condition_Type") == "Squad"]
    result_rows = [
        r for r in cond_rows
        if (
            (r.get("Condition_Type") == "Match_Result" and r.get("Target") in {"Win", "Clean Sheet"})
            or (r.get("Condition_Type") == "Match_Count" and r.get("Target") in {"Play", "Complete"})
        )
    ]

    lines = []
    for rr in result_rows:
        target = rr.get("Target", "")
        # Clean Sheet is already a direct 기타 수행 condition.
        if target == "Clean Sheet":
            lines.append(result_action_korean(target, int(rr.get("Count") or 0), rr.get("Per_Match", "No")))
            continue

        action_text = result_action_korean(target, int(rr.get("Count") or 0), rr.get("Per_Match", "No"))
        if squad_rows:
            for sr in squad_rows:
                req = korean_squad_requirement_line(sr)
                if req:
                    lines.append(f"{req}하고 {action_text}")
                else:
                    lines.append(action_text)
        else:
            lines.append(action_text)

    return unique_list(lines)


def compact_squad_requirements(cond_rows: List[Dict]) -> str:
    # Consolidate same Attribute/Value/Position, keep max Count.
    grouped = {}
    for r in cond_rows:
        if r["Condition_Type"] != "Squad":
            continue
        key = (r.get("Position", ""), r.get("Attribute", ""), r.get("Value", ""))
        grouped[key] = max(grouped.get(key, 0), int(r.get("Count") or 0))

    lines = []
    for (pos, attr, val), cnt in sorted(grouped.items(), key=lambda x: (x[0][0], x[0][1], x[0][2])):
        prefix = f"{pos}: " if pos else ""
        lines.append(f"{prefix}{val} ≥{cnt}")
    return "\n".join(lines) if lines else "-"


def compact_performer_requirements(cond_rows: List[Dict], target_filter: Optional[str] = None) -> str:
    """
    Compact performer rows.
    target_filter='Goal' collects both normal Goal and special goal types
    such as Outside Box Goal, Header Goal, Finesse Goal, Hat Trick.
    target_filter='Assist' collects both normal Assist and special assist types
    such as Through Ball Assist, Cross Assist, Incisive Pass Assist.
    """
    rows = [r for r in cond_rows if r["Condition_Type"] == "Performer"]
    if target_filter == "Goal":
        rows = [r for r in rows if is_goal_target(r.get("Target", ""))]
    elif target_filter == "Assist":
        rows = [r for r in rows if is_assist_target(r.get("Target", ""))]
    elif target_filter:
        rows = [r for r in rows if r["Target"] == target_filter]

    lines = []
    for r in rows:
        val = r["Value"]
        target = r["Target"]
        cnt = int(r.get("Count") or 0)
        if r.get("Per_Match") == "Yes":
            lines.append(f"{val}: {cnt}경기 각 {target} 1회")
        else:
            lines.append(f"{val}: {target} {cnt}회")
    return "\n".join(unique_list(lines)) if lines else "-"


def compact_other_actions(cond_rows: List[Dict]) -> str:
    """
    기타 수행/경기 내 특수 조건을 표시합니다.

    v30 기준:
    - 미션별 세부목표의 Win/Play/Complete/Clean Sheet도 반드시 To do(기타)에 표시합니다.
    - 선발 조건이 붙은 Win/Play는 '○○ 선수 n명 이상 선발하고 n승/플레이'로 연결해 표시합니다.
    - Goal/Assist 계열은 Goals/Assists로 보냅니다.
    """
    lines = []

    # First, include mission-level result objectives with squad context.
    lines.extend(mission_result_todo_lines(cond_rows))

    for r in cond_rows:
        target = r.get("Target", "")

        if r["Condition_Type"] == "Other_Action":
            # Raw fallback 문장 안에 goal/assist가 있더라도 파싱 불확실성을 보존하기 위해 유지합니다.
            lines.append(r["Operational_Note"])

        if (
            r["Condition_Type"] == "Performer"
            and not is_goal_target(target)
            and not is_assist_target(target)
        ):
            lines.append(r["Operational_Note"])

    return "\n".join(unique_list(lines)) if lines else "-"



def compact_results(cond_rows: List[Dict]) -> str:
    """
    Results에는 승리/플레이처럼 전체 경기 수 운영에 직접 관련되는 조건만 둡니다.
    Clean Sheet는 Other_Actions로 이동합니다.
    """
    lines = []
    for r in cond_rows:
        if r["Condition_Type"] == "Match_Result" and r["Target"] == "Win":
            lines.append(f"Win {r['Count']}")
        elif r["Condition_Type"] == "Match_Count":
            lines.append(f"{r['Target']} {r['Count']}")
    return "\n".join(unique_list(lines)) if lines else "-"


def min_matches_for_group(mission_rows: List[Dict], cond_rows: List[Dict]) -> int:
    candidates = []
    for m in mission_rows:
        action = m.get("Action", "")
        cnt = int(m.get("Action_Count") or 0)
        if action in {"Win", "Play", "Clean Sheet", "Complete"}:
            candidates.append(cnt)
        elif m.get("Per_Match") == "Yes":
            candidates.append(cnt)
    for c in cond_rows:
        if c.get("Per_Match") == "Yes":
            candidates.append(int(c.get("Count") or 0))
        if c["Condition_Type"] in {"Match_Result", "Match_Count"}:
            candidates.append(int(c.get("Count") or 0))
    return max(candidates) if candidates else 1



def compact_for_mission(cond_rows: List[Dict], mission_row: Dict) -> Dict:
    """Return one-row operational fields for a single mission."""
    squad = compact_squad_requirements(cond_rows)
    goals = compact_performer_requirements(cond_rows, "Goal")
    assists = compact_performer_requirements(cond_rows, "Assist")
    other_actions = compact_other_actions(cond_rows)
    results = compact_results(cond_rows)

    performer_vals = []
    for c in cond_rows:
        if c.get("Condition_Type") == "Performer" and c.get("Value") and c.get("Value") != "Any":
            performer_vals.append(c.get("Value"))

    # Need matches for this mission only.
    need_candidates = []
    if mission_row.get("Action") in {"Win", "Play", "Clean Sheet", "Complete"}:
        try:
            need_candidates.append(int(mission_row.get("Action_Count") or 0))
        except Exception:
            pass
    if mission_row.get("Per_Match") == "Yes":
        try:
            need_candidates.append(int(mission_row.get("Action_Count") or 0))
        except Exception:
            pass

    for c in cond_rows:
        if c.get("Per_Match") == "Yes":
            try:
                need_candidates.append(int(c.get("Count") or 0))
            except Exception:
                pass
        if c.get("Condition_Type") in {"Match_Result", "Match_Count"}:
            try:
                need_candidates.append(int(c.get("Count") or 0))
            except Exception:
                pass

    min_matches = max(need_candidates) if need_candidates else 1

    return {
        "Starting_XI": squad,
        "Performer_Candidates": "\n".join(unique_list(performer_vals)) if performer_vals else "-",
        "Goals": goals,
        "Assists": assists,
        "Other_Actions": other_actions,
        "Results": results,
        "Min_Matches": min_matches,
    }


def build_action_plan(df_mission: pd.DataFrame, df_conditions: pd.DataFrame) -> pd.DataFrame:
    """
    세부목표 1개당 1행.
    각 행에 선발조건, Goals, Assists, Other_Actions, Results를 분리 표시합니다.
    나중에 Completed=Yes인 Mission_ID를 제외하기 쉬운 구조입니다.
    """
    if df_mission is None or df_mission.empty:
        return pd.DataFrame()

    cond_map = defaultdict(list)
    if df_conditions is not None and not df_conditions.empty:
        for _, c in df_conditions.iterrows():
            cond_map[c["Mission_ID"]].append(c.to_dict())

    rows = []
    for _, m in df_mission.iterrows():
        mission = m.to_dict()
        conds = cond_map.get(mission["Mission_ID"], [])
        fields = compact_for_mission(conds, mission)

        rows.append({
            "Completed": "",
            "Is_New": mission.get("Is_New", "No"),
            "Mission_ID": mission.get("Mission_ID", ""),
            "Mode_Group": mission.get("Mode_Group", ""),
            "Campaign": mission.get("Campaign", ""),
            "Campaign_Expires": mission.get("Campaign_Expires", ""),
            "Campaign_Deadline_KST": mission.get("Campaign_Deadline_KST", ""),
            "Campaign_Deadline_ISO": mission.get("Campaign_Deadline_ISO", ""),
            "Mission_Name": mission.get("Mission_Name", ""),
            "Starting_XI": fields["Starting_XI"],
            "Performer_Candidates": fields["Performer_Candidates"],
            "Goals": fields["Goals"],
            "Assists": fields["Assists"],
            "Other_Actions": fields["Other_Actions"],
            "Results": fields["Results"],
            "Min_Matches": fields["Min_Matches"],
            "Per_Match": mission.get("Per_Match", ""),
            "Difficulty": mission.get("Difficulty", ""),
            "Reward": mission.get("Reward", ""),
            "Raw_Description": mission.get("Raw_Description", ""),
            "Source_URL": mission.get("Source_URL", ""),
        })

    return pd.DataFrame(rows)


def build_plan_summary_from_action_plan(df_action: pd.DataFrame) -> pd.DataFrame:
    """
    Action_Plan을 모드별로 압축한 요약.
    세부목표 추적은 Action_Plan에서 하고, Summary는 게임 시작 전 최종 확인용입니다.
    """
    if df_action is None or df_action.empty:
        return pd.DataFrame()

    rows = []
    for mode_group, group in df_action.groupby("Mode_Group", dropna=False):
        def join_unique(col):
            if col in {"Goals", "Assists", "Other_Actions"} and "Is_New" in group.columns:
                return unique_join_from_column_with_new(group, col)
            if col == "Starting_XI" and "Is_New" in group.columns:
                return unique_join_starting_with_new(group)
            if col == "Performer_Candidates" and "Is_New" in group.columns:
                return performer_candidates_excluding_starting_with_new(group)

            vals = []
            for v in group[col].dropna().astype(str).tolist():
                if not v or v == "-":
                    continue
                vals.extend([x.strip() for x in v.split("\n") if x.strip() and x.strip() != "-"])
            vals = unique_list(vals)
            if col == "Starting_XI":
                vals = consolidate_ge_lines(vals)
            if col == "Performer_Candidates":
                # Avoid repeating values already required in Starting_XI.
                starting_vals = get_starting_requirement_values("\n".join(consolidate_ge_lines([
                    x.strip()
                    for vv in group["Starting_XI"].dropna().astype(str).tolist()
                    for x in vv.split("\n")
                    if x.strip() and x.strip() != "-"
                ])))
                vals = [v for v in vals if v not in starting_vals]
            return "\n".join(vals) if vals else "-"

        rows.append({
            "Mode_Group": mode_group,
            "Plan": f"Plan {len(rows)+1}",
            "Min_Matches": int(group["Min_Matches"].max()) if "Min_Matches" in group and not group.empty else "",
            "Starting_XI": join_unique("Starting_XI"),
            "Performer_Candidates": join_unique("Performer_Candidates"),
            "Goals": join_unique("Goals"),
            "Assists": join_unique("Assists"),
            "Other_Actions": join_unique("Other_Actions"),
            "Results": "\n".join(consolidate_result_lines(group["Results"].dropna().astype(str).tolist())),
            "Covered_Campaigns": "\n".join(unique_list(group["Campaign"].dropna().astype(str).tolist())),
        })
    return pd.DataFrame(rows)


def build_mode_plan(df_mission: pd.DataFrame, df_conditions: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if df_mission.empty:
        return pd.DataFrame()

    for mode_group, group in df_mission.groupby("Mode_Group", dropna=False):
        mission_ids = set(group["Mission_ID"].tolist())
        conds = df_conditions[df_conditions["Mission_ID"].isin(mission_ids)].to_dict("records")
        missions = group.to_dict("records")

        goals = compact_performer_requirements(conds, "Goal")
        assists = compact_performer_requirements(conds, "Assist")
        other_actions = compact_other_actions(conds)
        other_actions = compact_other_actions(conds)

        # 실행계획에서는 하위목표명(Win 3, Score 2 등)보다 캠페인명이 더 유용함.
        # 같은 캠페인의 하위목표가 여러 개여도 1번만 표시.
        covered = unique_list(group["Campaign"].dropna().astype(str).tolist())

        rows.append({
            "Mode_Group": mode_group,
            "Plan": f"Plan {len(rows)+1}",
            "Min_Matches": min_matches_for_group(missions, conds),
            "Starting_XI": compact_squad_requirements(conds),
            "Bench_or_Performer_Candidates": compact_bench_candidates(conds),
            "Goals": goals,
            "Assists": assists,
            "Other_Actions": other_actions,
            "Results": compact_results(conds),
            "Covered_Missions": "\n".join(covered),
            "Conflict_Note": detect_conflict_note(conds),
        })
    return pd.DataFrame(rows)


def compact_bench_candidates(cond_rows: List[Dict]) -> str:
    vals = []
    for c in cond_rows:
        if c["Condition_Type"] == "Performer" and c["Value"] != "Any":
            vals.append(c["Value"])
    return "\n".join(unique_list(vals)) if vals else "-"


def detect_conflict_note(cond_rows: List[Dict]) -> str:
    # Basic rule: same exact position has multiple different required values.
    by_pos = defaultdict(set)
    for c in cond_rows:
        if c["Condition_Type"] == "Squad" and c.get("Position"):
            by_pos[c["Position"]].add(c["Value"])

    notes = []
    for pos, vals in by_pos.items():
        if len(vals) >= 2:
            notes.append(f"{pos} position has multiple conditions: {', '.join(sorted(vals))}. Same player 가능 여부 확인.")

    return "\n".join(notes) if notes else "-"


def build_conflict_check(df_plan: pd.DataFrame, df_conditions: pd.DataFrame) -> pd.DataFrame:
    rows = []

    # Mode groups themselves cannot be same match if different event/mode.
    if not df_plan.empty:
        for _, r in df_plan.iterrows():
            if "Event" in str(r["Mode_Group"]) or "event" in str(r["Mode_Group"]):
                rows.append({
                    "Conflict_Type": "Mode/Event separation",
                    "Mode_Group": r["Mode_Group"],
                    "Condition_A": "Event-specific mode",
                    "Condition_B": "Other modes",
                    "Suggested_Action": "별도 Plan으로 수행. 같은 판 병행 불가.",
                })

    # Same position conflicts
    if not df_conditions.empty:
        squad = df_conditions[df_conditions["Condition_Type"] == "Squad"]
        for pos, g in squad.groupby("Position", dropna=False):
            if not pos:
                continue
            vals = sorted(set(g["Value"].dropna().astype(str)))
            if len(vals) >= 2:
                rows.append({
                    "Conflict_Type": "Same-position attribute check",
                    "Mode_Group": "Check relevant missions",
                    "Condition_A": pos,
                    "Condition_B": " / ".join(vals),
                    "Suggested_Action": "한 선수가 모든 조건을 만족하면 병행, 아니면 스쿼드 분리.",
                })

    if not rows:
        rows.append({
            "Conflict_Type": "None detected by basic rules",
            "Mode_Group": "-",
            "Condition_A": "-",
            "Condition_B": "-",
            "Suggested_Action": "Raw 미션 조건은 별도 확인 권장.",
        })

    return pd.DataFrame(rows)



def one_line(text) -> str:
    """Convert multi-line plan cell to a compact comma-separated Korean-report line."""
    t = clean_text(str(text or ""))
    if not t or t == "-":
        return ""
    return t.replace("\n", ", ")


def build_korean_report(df_plan: pd.DataFrame, df_seasonal_plan: Optional[pd.DataFrame] = None) -> str:
    """
    Mode_Plan을 바탕으로 게임 중 바로 볼 수 있는 한글 실행 요약을 생성합니다.
    기본 보고서는 Seasonal을 제외한 df_plan 기준입니다.
    """
    lines = []
    lines.append("[오늘의 FUT.GG Objective 수행 요약]")
    lines.append("")
    lines.append("□ 핵심 원칙")
    lines.append("  - 아래 계획은 Seasonal을 제외한 기본 실행계획입니다.")
    lines.append("  - 선발 조건은 스쿼드에 넣기만 하면 되는 조건입니다.")
    lines.append("  - 득점·어시스트·기타 수행 조건은 해당 선수가 직접 수행해야 하는 조건입니다.")
    lines.append("")

    if df_plan is None or df_plan.empty:
        lines.append("□ 우선 수행 모드")
        lines.append("  - 정리할 실행계획이 없습니다.")
        lines.append("")
    else:
        lines.append("□ 우선 수행 모드")
        for idx, row in df_plan.reset_index(drop=True).iterrows():
            mode = row.get("Mode_Group", "")
            matches = row.get("Min_Matches", "")
            starting = one_line(row.get("Starting_XI", ""))
            candidates = one_line(row.get("Bench_or_Performer_Candidates", ""))
            goals = one_line(row.get("Goals", ""))
            assists = one_line(row.get("Assists", ""))
            other = one_line(row.get("Other_Actions", ""))
            results = one_line(row.get("Results", ""))
            covered = one_line(row.get("Covered_Missions", ""))
            conflict = one_line(row.get("Conflict_Note", ""))

            lines.append(f"  {idx + 1}) {mode}")
            lines.append(f"     - 최소 경기 수: {matches}경기")
            if starting:
                lines.append(f"     - 선발: {starting}")
            if candidates:
                lines.append(f"     - 교체/수행 후보: {candidates}")
            if goals:
                lines.append(f"     - 득점: {goals}")
            if assists:
                lines.append(f"     - 어시스트: {assists}")
            if other:
                lines.append(f"     - 기타 수행: {other}")
            if results:
                lines.append(f"     - 결과조건: {results}")
            if covered:
                lines.append(f"     - 포함 미션: {covered}")
            if conflict:
                lines.append(f"     - 주의: {conflict}")
            lines.append("")

    if df_seasonal_plan is not None and not df_seasonal_plan.empty:
        lines.append("□ Seasonal 미션 별도 참고")
        lines.append("  - Seasonal은 기본 실행계획에서 제외했습니다.")
        lines.append("  - 필요할 때만 아래 내용을 참고하세요.")
        for idx, row in df_seasonal_plan.reset_index(drop=True).iterrows():
            mode = row.get("Mode_Group", "")
            matches = row.get("Min_Matches", "")
            starting = one_line(row.get("Starting_XI", ""))
            goals = one_line(row.get("Goals", ""))
            assists = one_line(row.get("Assists", ""))
            results = one_line(row.get("Results", ""))

            lines.append(f"  {idx + 1}) {mode} / 최소 {matches}경기")
            detail_parts = []
            if starting:
                detail_parts.append(f"선발: {starting}")
            if goals:
                detail_parts.append(f"득점: {goals}")
            if assists:
                detail_parts.append(f"어시: {assists}")
            if results:
                detail_parts.append(f"결과: {results}")
            if detail_parts:
                lines.append("     - " + " / ".join(detail_parts))
        lines.append("")

    lines.append("□ 확인 필요")
    lines.append("  - Event 전용 미션은 다른 모드와 같은 판에서 병행할 수 없습니다.")
    lines.append("  - 같은 포지션에 서로 다른 선발 조건이 걸리면 한 선수가 동시에 충족 가능한지 확인하세요.")
    lines.append("  - 새 문장 유형은 파싱이 누락될 수 있으므로 애매한 조건은 Mission_DB의 Raw_Description을 확인하세요.")

    return "\n".join(lines)


def korean_report_dataframe(report_text: str) -> pd.DataFrame:
    return pd.DataFrame({"한글 실행 요약": report_text.split("\n")})


def build_parse_check(df_mission: pd.DataFrame, df_conditions: pd.DataFrame) -> pd.DataFrame:
    """
    Flag missions whose raw text contains special-action keywords.
    Useful for checking whether new FUT.GG wording was parsed correctly.
    """
    if df_mission is None or df_mission.empty:
        return pd.DataFrame()

    rows = []
    special_keywords = [
        ("outside box/from distance", r"outside\s+the\s+box|from\s+distance|long[- ]range", "Outside Box Goal"),
        ("finesse", r"\bfinesse\b", "Finesse Goal"),
        ("low driven", r"low\s+driven", "Low Driven Goal"),
        ("through ball", r"through\s+ball", "Through Ball Assist"),
        ("header", r"\bheader(s)?\b|headed\s+goals?", "Header Goal"),
        ("volley", r"\bvolley(s)?\b", "Volley Goal"),
        ("tackle", r"\btackles?\b", "Tackle"),
        ("interception", r"\binterceptions?\b", "Interception"),
        ("concede limit", r"concede\s+no\s+more\s+than|concede\s+(?:less|fewer)\s+than", "Concede Limit"),
        ("goal and assist", r"score\s+and\s+assist|goal\s+and\s+assist|goals?\s*&\s*assists?", "Goal+Assist Same Match"),
        ("perfect hat trick", r"perfect\s+hat[- ]?trick", "Perfect Hat Trick"),
        ("hat trick", r"hat[- ]?trick", "Hat Trick"),
        ("weak foot", r"weak\s+foot", "Weak Foot"),
        ("skill moves", r"skill\s+moves?", "Skill Moves"),
        ("power shot", r"power\s+shot", "Power Shot Goal"),
        ("chip shot", r"chip\s+shot", "Chip Shot Goal"),
    ]

    cond_by_mid = defaultdict(list)
    if df_conditions is not None and not df_conditions.empty:
        for _, c in df_conditions.iterrows():
            cond_by_mid[c["Mission_ID"]].append(str(c.get("Target", "")))

    for _, m in df_mission.iterrows():
        raw = str(m.get("Raw_Description", ""))
        mid = m.get("Mission_ID", "")
        parsed_targets = set(cond_by_mid.get(mid, []))
        for label, pat, expected_target in special_keywords:
            if re.search(pat, raw, re.I):
                rows.append({
                    "Mission_ID": mid,
                    "Campaign": m.get("Campaign", ""),
                    "Mission_Name": m.get("Mission_Name", ""),
                    "Keyword": label,
                    "Expected_Target": expected_target,
                    "Parsed_Targets": ", ".join(sorted(parsed_targets)),
                    "Check_Result": "OK" if expected_target in parsed_targets else "CHECK",
                    "Raw_Description": raw,
                })

    # Broad safety net: if the raw sentence contains unusual action words but only generic
    # Goal/Assist/Other was parsed, flag it for review.
    broad_suspicious = [
        "outside", "distance", "finesse", "low driven", "through ball", "header", "volley",
        "tackle", "interception", "concede", "hat-trick", "hat trick", "weak foot",
        "skill moves", "power shot", "chip shot", "goal and assist", "score and assist"
    ]
    existing_pairs = {(r.get("Mission_ID"), r.get("Keyword")) for r in rows}

    for _, m in df_mission.iterrows():
        raw = str(m.get("Raw_Description", ""))
        mid = m.get("Mission_ID", "")
        parsed_targets = set(cond_by_mid.get(mid, []))
        hits = [kw for kw in broad_suspicious if kw.lower() in raw.lower()]
        if hits and not any(t not in {"Goal", "Assist", "Win", "Play", "Clean Sheet", "Other"} for t in parsed_targets):
            key = (mid, "broad suspicious")
            if key not in existing_pairs:
                rows.append({
                    "Mission_ID": mid,
                    "Campaign": m.get("Campaign", ""),
                    "Mission_Name": m.get("Mission_Name", ""),
                    "Keyword": "broad suspicious: " + ", ".join(hits),
                    "Expected_Target": "Review needed",
                    "Parsed_Targets": ", ".join(sorted(parsed_targets)),
                    "Check_Result": "CHECK",
                    "Raw_Description": raw,
                })

    return pd.DataFrame(rows)



TARGET_REPORT_MODE = "Squad Battles / Rivals / Champions / Live Events"


def split_cell_lines(value) -> List[str]:
    text = clean_text(str(value or ""))
    if not text or text == "-":
        return []
    return [clean_text(x) for x in str(value).split("\n") if clean_text(x) and clean_text(x) != "-"]


def unique_join_from_column(df: pd.DataFrame, col: str) -> str:
    return unique_join_from_column_consolidated(df, col, consolidate_ge=False)



def parse_ge_requirement_line(line: str):
    """
    Parse simple requirement lines like:
      Scotland ≥2
      CONMEBOL Libertadores ≥1
      LB: Bundesliga ≥1
    Returns (key, label, count) or None.
    """
    t = clean_text(line)
    if not t or t == "-":
        return None

    m = re.match(r"^(?:(?P<pos>[A-Z]{1,3}):\s*)?(?P<label>.+?)\s*≥\s*(?P<count>[0-9]+)\s*$", t)
    if not m:
        return None

    pos = clean_text(m.group("pos") or "")
    label = clean_text(m.group("label") or "")
    count = int(m.group("count"))
    key = (pos, label)
    display_label = f"{pos}: {label}" if pos else label
    return key, display_label, count


def consolidate_ge_lines(lines: List[str]) -> List[str]:
    """
    Same requirement with different counts should be collapsed to the largest count.
      Scotland ≥1 + Scotland ≥2 -> Scotland ≥2
      Uruguay ≥2 + Uruguay ≥1 -> Uruguay ≥2
    Non-matching lines are kept once.
    """
    max_by_key = {}
    label_by_key = {}
    passthrough = []

    for line in lines:
        parsed = parse_ge_requirement_line(line)
        if parsed:
            key, label, count = parsed
            max_by_key[key] = max(max_by_key.get(key, 0), count)
            label_by_key[key] = label
        else:
            passthrough.append(line)

    out = []
    for key in sorted(max_by_key.keys(), key=lambda x: (x[0], x[1])):
        out.append(f"{label_by_key[key]} ≥{max_by_key[key]}")

    out.extend(unique_list(passthrough))
    return out


def get_starting_requirement_values(starting_text: str) -> set:
    """
    Return labels already required in Starting_XI.
    Used to avoid repeating them in performer-candidate summary.
    """
    vals = set()
    for line in split_cell_lines(starting_text):
        parsed = parse_ge_requirement_line(line)
        if parsed:
            _, label, _ = parsed
            # remove optional position prefix from comparison target
            label = re.sub(r"^[A-Z]{1,3}:\s*", "", label)
            vals.add(clean_text(label))
        else:
            vals.add(clean_text(line))
    return vals


def unique_join_from_column_consolidated(df: pd.DataFrame, col: str, consolidate_ge: bool = False) -> str:
    if df is None or df.empty or col not in df.columns:
        return "-"
    vals = []
    for v in df[col].dropna().astype(str).tolist():
        vals.extend(split_cell_lines(v))
    vals = unique_list(vals)
    if consolidate_ge:
        vals = consolidate_ge_lines(vals)
    return "\n".join(vals) if vals else "-"


def performer_candidates_excluding_starting(df: pd.DataFrame) -> str:
    """
    SB_Report용.
    Starting_XI에서 이미 요구되는 값은 수행 후보에서 빼고,
    진짜 추가로 준비할 후보만 보여줍니다.
    """
    if df is None or df.empty:
        return "-"

    starting_text = unique_join_from_column_consolidated(df, "Starting_XI", consolidate_ge=True)
    starting_values = get_starting_requirement_values(starting_text)

    vals = []
    if "Performer_Candidates" in df.columns:
        for v in df["Performer_Candidates"].dropna().astype(str).tolist():
            vals.extend(split_cell_lines(v))

    out = []
    for v in unique_list(vals):
        if v not in starting_values:
            out.append(v)

    return "\n".join(out) if out else "-"



def consolidate_result_lines(text_or_lines) -> List[str]:
    """
    Consolidate result requirements.
    Example:
      Win 2, Win 7 -> Win 7
      Play 10, Play 5 -> Play 10
    """
    if isinstance(text_or_lines, str):
        lines = split_cell_lines(text_or_lines)
    else:
        lines = []
        for v in text_or_lines:
            lines.extend(split_cell_lines(v))

    max_by_action = {}
    passthrough = []

    for line in lines:
        t = clean_text(line)
        m = re.match(r"^(Win|Play|Complete)\s+([0-9]+)$", t, re.I)
        if m:
            action = m.group(1).capitalize()
            count = int(m.group(2))
            max_by_action[action] = max(max_by_action.get(action, 0), count)
        else:
            # Keep nonstandard result lines once.
            passthrough.append(t)

    order = ["Win", "Play", "Complete"]
    out = [f"{a} {max_by_action[a]}" for a in order if a in max_by_action]
    out.extend(unique_list(passthrough))
    return out


def result_summary_for_report(df: pd.DataFrame) -> str:
    if df is None or df.empty or "Results" not in df.columns:
        return "-"
    return "\n".join(consolidate_result_lines(df["Results"].dropna().astype(str).tolist())) or "-"



def unique_join_starting_with_new(df: pd.DataFrame, marker: str = " (신규미션)") -> str:
    """
    Starting_XI summary with conservative new markers:
    - If a requirement is only from new missions, mark it.
    - If new missions increase the required count, mark it.
    - If the same requirement already existed at the same or higher count, no marker.
    """
    if df is None or df.empty or "Starting_XI" not in df.columns:
        return "-"

    old_max = {}
    all_max = {}
    label_by_key = {}

    for _, row in df.iterrows():
        is_new = str(row.get("Is_New", "No")).strip().lower() == "yes"
        for line in split_cell_lines(row.get("Starting_XI", "")):
            parsed = parse_ge_requirement_line(line)
            if not parsed:
                key = ("RAW", line)
                label = line
                cnt = 1
            else:
                key, label, cnt = parsed

            all_max[key] = max(all_max.get(key, 0), cnt)
            label_by_key[key] = label
            if not is_new:
                old_max[key] = max(old_max.get(key, 0), cnt)

    out = []
    for key in sorted(all_max.keys(), key=lambda x: str(x)):
        label = label_by_key[key]
        cnt = all_max[key]
        if key[0] == "RAW":
            line = label
        else:
            line = f"{label} ≥{cnt}"

        if cnt > old_max.get(key, 0) and marker not in line:
            line += marker
        out.append(line)

    return "\n".join(out) if out else "-"


def performer_candidates_excluding_starting_with_new(df: pd.DataFrame, marker: str = " (신규미션)") -> str:
    """
    Performer candidates summary with conservative new markers:
    - Remove values already required in Starting_XI.
    - Mark only candidates that appear in new rows and did not appear in old rows.
    """
    if df is None or df.empty:
        return "-"

    starting_text = unique_join_starting_with_new(df, marker="")
    starting_values = get_starting_requirement_values(starting_text)

    old_vals = []
    new_vals = []

    if "Performer_Candidates" in df.columns:
        for _, row in df.iterrows():
            vals = split_cell_lines(row.get("Performer_Candidates", ""))
            vals = [v for v in vals if v not in starting_values]
            if str(row.get("Is_New", "No")).strip().lower() == "yes":
                new_vals.extend(vals)
            else:
                old_vals.extend(vals)

    old_set = set(unique_list(old_vals))
    new_set = set(unique_list(new_vals))

    out = []
    for v in unique_list(old_vals + new_vals):
        if v in new_set and v not in old_set:
            out.append(f"{v}{marker}")
        else:
            out.append(v)

    return "\n".join(out) if out else "-"


def campaigns_for_column(df: pd.DataFrame, cols: List[str]) -> str:
    """
    SB_Report C column helper.
    Shows campaigns that contribute to a given section. This is for report/poster
    context, not for gameplay instruction.
    """
    if df is None or df.empty:
        return "-"

    rows = []
    for _, row in df.iterrows():
        has_value = False
        for col in cols:
            if col in df.columns:
                vals = split_cell_lines(row.get(col, ""))
                if vals:
                    has_value = True
                    break
        if has_value:
            rows.append(campaign_display_from_row(row))

    vals = unique_list(rows)
    return "\n".join(vals) if vals else "-"



ANY_ULTIMATE_MODE = "Any Ultimate Team mode"


def is_any_ultimate_mode(mode: str) -> bool:
    return clean_text(str(mode or "")).lower() == ANY_ULTIMATE_MODE.lower()


def select_report_df(
    df_action: pd.DataFrame,
    target_mode: str,
    include_any_ultimate: bool = True,
    only_new: bool = False,
) -> pd.DataFrame:
    """
    Select rows for a mode-specific report.

    Important planning rule:
    - 'Any Ultimate Team mode' objectives can be completed together with SB/Rivals/Champions/Live Events
      or event-specific modes, so include them in every non-Any mode report.
    - If only_new=True, filter after adding the Any rows, so newly added Any-mode missions also show
      in each relevant new-mission report.
    """
    if df_action is None or df_action.empty or "Mode_Group" not in df_action.columns:
        return pd.DataFrame()

    target = clean_text(str(target_mode or ""))
    mode_series = df_action["Mode_Group"].astype(str).map(clean_text)

    if include_any_ultimate and target and not is_any_ultimate_mode(target):
        mask = mode_series.eq(target) | mode_series.map(is_any_ultimate_mode)
    else:
        mask = mode_series.eq(target)

    df = df_action[mask].copy()

    if only_new and "Is_New" in df.columns:
        df = df[df["Is_New"].astype(str).str.lower().eq("yes")].copy()

    return df


def report_mode_targets(df_action: pd.DataFrame, include_sb: bool = False, only_new: bool = False) -> List[str]:
    """
    Return mode groups that should receive their own report sheet.
    - Exclude Any Ultimate Team mode if there are other modes, because it is included as common rows.
    - SB_Report already covers TARGET_REPORT_MODE, so Mode_Report normally excludes it.
    - New_Report can include TARGET_REPORT_MODE because it is a separate 신규미션 view.
    """
    if df_action is None or df_action.empty or "Mode_Group" not in df_action.columns:
        return []

    df = df_action.copy()
    if only_new and "Is_New" in df.columns:
        df = df[df["Is_New"].astype(str).str.lower().eq("yes")].copy()

    modes = unique_list(df["Mode_Group"].dropna().astype(str).map(clean_text).tolist())
    non_any = [m for m in modes if not is_any_ultimate_mode(m)]

    if non_any:
        modes = non_any

    if not include_sb:
        modes = [m for m in modes if m != TARGET_REPORT_MODE]

    return sorted(modes)


def build_report_index(df_action: pd.DataFrame, sheet_map: Dict[str, str], only_new: bool = False) -> pd.DataFrame:
    """
    Index sheet for Mode_Report/New_Report sheets.
    Count is based on the actual report selection, including Any Ultimate Team common rows.
    """
    rows = []
    modes = list(sheet_map.keys())

    if not modes:
        return pd.DataFrame([{
            "Sheet": "-",
            "Mode_Group": "-",
            "Mission_Count": 0,
            "Min_Matches": "-",
            "Includes_Any_Ultimate": "-",
            "Related_Campaigns": "신규 미션 없음" if only_new else "정리할 모드 없음",
        }])

    for mode in modes:
        df = select_report_df(df_action, mode, include_any_ultimate=True, only_new=only_new)
        any_included = False
        if not df.empty and "Mode_Group" in df.columns:
            any_included = any(is_any_ultimate_mode(m) for m in df["Mode_Group"].dropna().astype(str))

        rows.append({
            "Sheet": sheet_map.get(mode, ""),
            "Mode_Group": mode,
            "Mission_Count": int(df["Mission_ID"].nunique()) if "Mission_ID" in df.columns and not df.empty else 0,
            "Min_Matches": int(df["Min_Matches"].max()) if "Min_Matches" in df.columns and not df.empty else "",
            "Includes_Any_Ultimate": "Yes" if any_included and not is_any_ultimate_mode(mode) else "No",
            "Related_Campaigns": "\n".join(unique_list(df["Campaign"].dropna().astype(str).tolist())) if "Campaign" in df.columns and not df.empty else "-",
        })

    return pd.DataFrame(rows)



CORE_STACK_MODE = TARGET_REPORT_MODE


def infer_play_context(mode_group: str, campaign: str = "", raw_description: str = "") -> str:
    """
    Infer the practical play context for a newly added specific-mode mission.

    Examples:
      - Ligue 1 TOTS Exhibition -> Live Events
      - TOTS: Champions Objective -> Champions
    """
    mode = clean_text(str(mode_group or ""))
    blob = f"{mode} {campaign} {raw_description}".lower()

    if is_any_ultimate_mode(mode) or mode == CORE_STACK_MODE:
        return ""

    if "champions" in blob:
        return "Champions"

    # FUT.GG event/exhibition/cup objectives are practically Live Events.
    if (
        "live event" in blob
        or "live events" in blob
        or "exhibition" in blob
        or re.search(r"\bcup\b", blob)
        or re.search(r"\bevent\b", blob)
    ):
        return "Live Events"

    return mode


def build_combo_targets(df_action: pd.DataFrame) -> List[Dict]:
    """
    Build combo-report targets only when new missions exist.

    A combo report answers:
      "If a new mission belongs to Champions or a Live Event, what can I complete
       together in that same play context?"

    It includes:
      - Any Ultimate Team mode rows
      - core SB/Rivals/Champions/Live Events rows
      - the new target mode rows, e.g. Ligue 1 TOTS Exhibition
      - generic context rows if present, e.g. Champions
    """
    if df_action is None or df_action.empty or "Is_New" not in df_action.columns:
        return []

    new_df = df_action[df_action["Is_New"].astype(str).str.lower().eq("yes")].copy()
    if new_df.empty:
        return []

    targets = {}
    for _, row in new_df.iterrows():
        mode = clean_text(str(row.get("Mode_Group", "")))
        if not mode or mode == CORE_STACK_MODE or is_any_ultimate_mode(mode):
            continue

        campaign = clean_text(str(row.get("Campaign", "")))
        raw = clean_text(str(row.get("Raw_Description", "")))
        context = infer_play_context(mode, campaign, raw)
        if not context:
            continue

        # One report per new specific mode/context combination.
        # If the mode is simply "Champions", key by context only.
        trigger_mode = mode
        key = (context, trigger_mode)
        item = targets.setdefault(key, {
            "Play_Context": context,
            "Trigger_Mode": trigger_mode,
            "New_Campaigns": [],
        })
        if campaign:
            item["New_Campaigns"].append(campaign)

    out = []
    for item in targets.values():
        item["New_Campaigns"] = "\n".join(unique_list(item.get("New_Campaigns", [])))
        out.append(item)

    return sorted(out, key=lambda x: (x["Play_Context"], x["Trigger_Mode"]))


def select_combo_report_df(df_action: pd.DataFrame, play_context: str, trigger_mode: str) -> pd.DataFrame:
    """
    Select rows that can be stacked in a newly triggered play context.

    Included mode groups:
      - Any Ultimate Team mode
      - Squad Battles / Rivals / Champions / Live Events core group
      - trigger_mode itself, e.g. Ligue 1 TOTS Exhibition
      - play_context itself, e.g. Champions or Live Events, if present
    """
    if df_action is None or df_action.empty or "Mode_Group" not in df_action.columns:
        return pd.DataFrame()

    context = clean_text(str(play_context or ""))
    trigger = clean_text(str(trigger_mode or ""))

    mode_series = df_action["Mode_Group"].astype(str).map(clean_text)

    allowed = {CORE_STACK_MODE, ANY_ULTIMATE_MODE}
    if trigger:
        allowed.add(trigger)
    if context:
        allowed.add(context)

    mask = mode_series.isin(allowed)

    df = df_action[mask].copy()

    # Keep original order from Action_Plan for readability.
    return df


def build_combo_report_rows(df_action: pd.DataFrame, target: Dict) -> pd.DataFrame:
    """
    Build a practical combined report for new specific-mode missions.
    """
    context = clean_text(str(target.get("Play_Context", "")))
    trigger = clean_text(str(target.get("Trigger_Mode", "")))

    display = trigger if trigger and trigger != context else context
    if context and trigger and trigger != context:
        display = f"{context} 기준: {trigger}"

    df = select_combo_report_df(df_action, play_context=context, trigger_mode=trigger)

    if df.empty:
        return pd.DataFrame({
            "구분": ["통합 수행 모드", "요약"],
            "내용": [display, "정리할 세부목표가 없습니다."],
            "관련 캠페인": ["-", "-"],
        })

    included_modes = unique_list(df["Mode_Group"].dropna().astype(str).map(clean_text).tolist())
    min_matches = int(df["Min_Matches"].max()) if "Min_Matches" in df.columns and not df.empty else ""

    rows = [
        {"구분": "통합 수행 모드", "내용": display, "관련 캠페인": "-"},
        {"구분": "포함 기준", "내용": " / ".join(included_modes), "관련 캠페인": "-"},
        {"구분": "최소 필요경기", "내용": f"{min_matches}경기", "관련 캠페인": "-"},
    ]

    result_lines = consolidate_result_lines(df["Results"].dropna().astype(str).tolist()) if "Results" in df.columns else []
    append_report_section(
        rows,
        "플레이 요건",
        [{"내용": line, "관련 캠페인": campaigns_for_column(df[df["Results"].astype(str).str.contains(re.escape(line), na=False)] if "Results" in df.columns else df, ["Results"])} for line in result_lines],
        empty_text="-",
    )

    append_report_section(rows, "필수 선발 요건", report_starting_rows(df))
    append_report_section(rows, "교체로 투입해도 가능", report_performer_candidate_rows(df))
    append_report_section(rows, "To do (득점)", report_lines_for_column(df, "Goals"))
    append_report_section(rows, "To do (어시스트)", report_lines_for_column(df, "Assists"))
    append_report_section(rows, "To do (기타)", report_lines_for_column(df, "Other_Actions"))

    return pd.DataFrame(rows)


def build_combo_report_index(df_action: pd.DataFrame, combo_targets: List[Dict], sheet_map: Dict[tuple, str]) -> pd.DataFrame:
    if not combo_targets:
        return pd.DataFrame([{
            "Sheet": "-",
            "Play_Context": "-",
            "Trigger_Mode": "-",
            "Mission_Count": 0,
            "Min_Matches": "-",
            "Included_Mode_Groups": "신규 특수 모드 없음",
            "New_Campaigns": "-",
        }])

    rows = []
    for target in combo_targets:
        context = target.get("Play_Context", "")
        trigger = target.get("Trigger_Mode", "")
        df = select_combo_report_df(df_action, play_context=context, trigger_mode=trigger)
        rows.append({
            "Sheet": sheet_map.get((context, trigger), ""),
            "Play_Context": context,
            "Trigger_Mode": trigger,
            "Mission_Count": int(df["Mission_ID"].nunique()) if "Mission_ID" in df.columns and not df.empty else 0,
            "Min_Matches": int(df["Min_Matches"].max()) if "Min_Matches" in df.columns and not df.empty else "",
            "Included_Mode_Groups": "\n".join(unique_list(df["Mode_Group"].dropna().astype(str).map(clean_text).tolist())) if "Mode_Group" in df.columns and not df.empty else "-",
            "New_Campaigns": target.get("New_Campaigns", "-"),
        })

    return pd.DataFrame(rows)


def build_mode_report_rows(
    df_action: pd.DataFrame,
    target_mode: str = TARGET_REPORT_MODE,
    only_new: bool = False,
    include_any_ultimate: bool = True,
) -> pd.DataFrame:
    """
    보고서용 시트.
    B열 내용과 C열 관련 캠페인이 1:1로 대응되도록 항목 1개당 1행으로 펼칩니다.

    include_any_ultimate=True이면 Any Ultimate Team mode 세부목표를 각 모드별 보고서에도 포함합니다.
    """
    report_label = "신규미션 대상 모드" if only_new else "대상 모드"

    if df_action is None or df_action.empty:
        return pd.DataFrame({
            "구분": [report_label, "요약"],
            "내용": [target_mode, "정리할 미션이 없습니다."],
            "관련 캠페인": ["-", "-"],
        })

    df = select_report_df(
        df_action,
        target_mode=target_mode,
        include_any_ultimate=include_any_ultimate,
        only_new=only_new,
    )

    if df.empty:
        msg = "해당 모드의 신규 미션이 없습니다." if only_new else "해당 모드에서 수행할 미션이 없습니다."
        return pd.DataFrame({
            "구분": [report_label, "요약"],
            "내용": [target_mode, msg],
            "관련 캠페인": ["-", "-"],
        })

    min_matches = int(df["Min_Matches"].max()) if "Min_Matches" in df.columns and not df.empty else ""
    any_included = (
        include_any_ultimate
        and not is_any_ultimate_mode(target_mode)
        and "Mode_Group" in df.columns
        and any(is_any_ultimate_mode(m) for m in df["Mode_Group"].dropna().astype(str))
    )

    rows = [
        {"구분": report_label, "내용": target_mode, "관련 캠페인": "-"},
    ]
    if any_included:
        rows.append({"구분": "공통 포함", "내용": "모든 Ultimate Team 모드 세부목표 포함", "관련 캠페인": "Any Ultimate Team mode"})
    rows.append({
        "구분": "최소 필요경기",
        "내용": f"{min_matches}경기",
        "관련 캠페인": "-",
    })

    # Result requirements are summary values, not always one-campaign-only.
    result_lines = consolidate_result_lines(df["Results"].dropna().astype(str).tolist()) if "Results" in df.columns else []
    append_report_section(
        rows,
        "플레이 요건",
        [{"내용": line, "관련 캠페인": campaigns_for_column(df[df["Results"].astype(str).str.contains(re.escape(line), na=False)] if "Results" in df.columns else df, ["Results"])} for line in result_lines],
        empty_text="-",
    )

    append_report_section(rows, "필수 선발 요건", report_starting_rows(df))
    append_report_section(rows, "교체로 투입해도 가능", report_performer_candidate_rows(df))
    append_report_section(rows, "To do (득점)", report_lines_for_column(df, "Goals"))
    append_report_section(rows, "To do (어시스트)", report_lines_for_column(df, "Assists"))
    append_report_section(rows, "To do (기타)", report_lines_for_column(df, "Other_Actions"))

    return pd.DataFrame(rows)


def build_mode_action_detail(df_action: pd.DataFrame, target_mode: str = TARGET_REPORT_MODE) -> pd.DataFrame:
    """
    보고서용 상세 근거. Action_Plan 전체보다 좁혀서, 대상 모드 행만 간결하게 보여줍니다.
    """
    if df_action is None or df_action.empty:
        return pd.DataFrame()

    df = select_report_df(df_action, target_mode=target_mode, include_any_ultimate=True, only_new=False)
    if df.empty:
        return pd.DataFrame()

    keep_cols = [
        "Completed", "Is_New", "Campaign", "Campaign_Expires", "Campaign_Deadline_KST", "Mission_Name", "Starting_XI", "Performer_Candidates",
        "Goals", "Assists", "Other_Actions", "Results", "Min_Matches", "Per_Match", "Difficulty"
    ]
    keep_cols = [c for c in keep_cols if c in df.columns]
    return df[keep_cols].copy()



def safe_excel_sheet_name(name: str, used: Optional[set] = None, max_len: int = 31) -> str:
    """
    Create a valid Excel sheet name.
    """
    used = used or set()
    s = clean_text(str(name or "Sheet"))
    s = re.sub(r"[\[\]\:\*\?\/\\]", "_", s)
    s = s[:max_len].strip() or "Sheet"

    base = s
    idx = 1
    while s in used:
        suffix = f"_{idx}"
        s = (base[: max_len - len(suffix)] + suffix).strip()
        idx += 1
    used.add(s)
    return s


def build_new_report_index(df_action: pd.DataFrame, sheet_map: Dict[str, str]) -> pd.DataFrame:
    """
    Index sheet for per-mode New_Report sheets.
    """
    if df_action is None or df_action.empty or "Is_New" not in df_action.columns:
        return pd.DataFrame([{
            "Sheet": "-",
            "Mode_Group": "-",
            "New_Mission_Count": 0,
            "Min_Matches": "-",
            "Related_Campaigns": "-",
        }])

    rows = []
    new_df = df_action[df_action["Is_New"].astype(str).str.lower().eq("yes")].copy()
    if new_df.empty:
        return pd.DataFrame([{
            "Sheet": "-",
            "Mode_Group": "-",
            "New_Mission_Count": 0,
            "Min_Matches": "-",
            "Related_Campaigns": "신규 미션 없음",
        }])

    for mode, group in new_df.groupby("Mode_Group", dropna=False):
        mode = str(mode)
        rows.append({
            "Sheet": sheet_map.get(mode, ""),
            "Mode_Group": mode,
            "New_Mission_Count": int(group["Mission_ID"].nunique()) if "Mission_ID" in group.columns else len(group),
            "Min_Matches": int(group["Min_Matches"].max()) if "Min_Matches" in group.columns and not group.empty else "",
            "Related_Campaigns": "\n".join(unique_list(group["Campaign"].dropna().astype(str).tolist())),
        })

    return pd.DataFrame(rows)


def write_excel(out_path: Path, df_mission: pd.DataFrame, df_conditions: pd.DataFrame, df_plan: pd.DataFrame,
                df_conflict: pd.DataFrame, df_links: pd.DataFrame, errors: List[Dict],
                df_seasonal_plan: Optional[pd.DataFrame] = None,
                df_plan_all: Optional[pd.DataFrame] = None,
                df_action: Optional[pd.DataFrame] = None,
                df_action_seasonal: Optional[pd.DataFrame] = None,
                df_action_all: Optional[pd.DataFrame] = None):
    df_errors = pd.DataFrame(errors)
    df_parse_check = build_parse_check(df_mission, df_conditions)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        if df_action is not None and not df_action.empty:
            build_mode_report_rows(df_action).to_excel(writer, sheet_name="SB_Report", index=False)
            sb_detail = build_mode_action_detail(df_action)
            if sb_detail is not None and not sb_detail.empty:
                sb_detail.to_excel(writer, sheet_name="SB_Action_Detail", index=False)

            used_sheet_names = set(writer.book.sheetnames)

            # 신규 특수 모드 통합 리포트:
            # 새로 나온 Ligue 1 TOTS Exhibition, TOTS: Champions Objective 같은 미션을 기준으로
            # Any Ultimate Team mode + 기본 SB/Rivals/Champions/Live Events + 해당 모드에서
            # 한꺼번에 중첩 가능한 세부목표를 정리합니다.
            combo_targets = build_combo_targets(df_action)
            combo_sheet_map = {}
            for idx, target in enumerate(combo_targets, start=1):
                sheet_name = safe_excel_sheet_name(f"Combo_Report_{idx:02d}", used_sheet_names)
                key = (target.get("Play_Context", ""), target.get("Trigger_Mode", ""))
                combo_sheet_map[key] = sheet_name
                build_combo_report_rows(df_action, target).to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )
            build_combo_report_index(df_action, combo_targets, combo_sheet_map).to_excel(writer, sheet_name="Combo_Report_Index", index=False)

            # 신규미션 전용: 모드별로 SB_Report와 같은 형식의 보고서 시트를 추가합니다.
            # Any Ultimate Team mode 신규 세부목표는 각 모드별 New_Report에도 함께 포함합니다.
            new_sheet_map = {}
            if "Is_New" in df_action.columns:
                new_targets = report_mode_targets(df_action, include_sb=True, only_new=True)
                for idx, mode in enumerate(new_targets, start=1):
                    sheet_name = safe_excel_sheet_name(f"New_Report_{idx:02d}", used_sheet_names)
                    new_sheet_map[mode] = sheet_name
                    build_mode_report_rows(df_action, target_mode=mode, only_new=True, include_any_ultimate=True).to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=False,
                    )

            build_report_index(df_action, new_sheet_map, only_new=True).to_excel(writer, sheet_name="New_Report_Index", index=False)

            df_action.to_excel(writer, sheet_name="Action_Plan", index=False)
            build_plan_summary_from_action_plan(df_action).to_excel(writer, sheet_name="Plan_Summary", index=False)

        df_plan.to_excel(writer, sheet_name="Mode_Plan", index=False)

        # 참고용: Seasonal만 따로, 그리고 전체 포함 실행계획도 별도 보관
        if df_action_seasonal is not None and not df_action_seasonal.empty:
            df_action_seasonal.to_excel(writer, sheet_name="Seasonal_Action_Plan", index=False)
        if df_action_all is not None and not df_action_all.empty:
            df_action_all.to_excel(writer, sheet_name="Action_Plan_All", index=False)
        if df_seasonal_plan is not None and not df_seasonal_plan.empty:
            df_seasonal_plan.to_excel(writer, sheet_name="Seasonal_Mode_Plan", index=False)
        if df_plan_all is not None and not df_plan_all.empty:
            df_plan_all.to_excel(writer, sheet_name="Mode_Plan_All", index=False)

        df_mission.to_excel(writer, sheet_name="Mission_DB", index=False)
        df_conditions.to_excel(writer, sheet_name="Condition_DB", index=False)
        df_conflict.to_excel(writer, sheet_name="Conflict_Check", index=False)
        if not df_parse_check.empty:
            df_parse_check.to_excel(writer, sheet_name="Parse_Check", index=False)
        df_links.to_excel(writer, sheet_name="Collected_URLs", index=False)
        df_errors.to_excel(writer, sheet_name="Errors", index=False)

        guide = pd.DataFrame([
            {"Item": "가장 먼저 볼 시트", "Description": "SB_Report: 대상 모드 기준 핵심 수행 요약입니다. B열 내용 1개와 C열 관련 캠페인이 1:1로 대응되도록 펼쳐 표시합니다."},
            {"Item": "SB_Action_Detail", "Description": "대상 모드의 세부목표별 근거표입니다. 완료 체크와 상세 확인용입니다."},
            {"Item": "Combo_Report_Index", "Description": "새로 나온 특수 모드 미션을 기준으로, 같은 모드에서 한꺼번에 중첩 가능한 세부목표를 묶은 Combo_Report 시트를 안내합니다."},
            {"Item": "Combo_Report_XX", "Description": "Any Ultimate Team mode + 기본 SB/Rivals/Champions/Live Events + 해당 신규 모드 세부목표를 함께 묶은 실전용 통합 리포트입니다."},
            {"Item": "New_Report_Index", "Description": "신규미션이 있는 모드별 New_Report 시트를 안내합니다. Any Ultimate Team mode 신규 세부목표도 함께 포함됩니다."},
            {"Item": "New_Report_XX", "Description": "신규미션만 대상으로 모드별 실행 요약을 SB_Report와 같은 형식으로 보여줍니다."},
            {"Item": "Action_Plan", "Description": "세부목표 1개당 1행으로 Goals/Assists/Other_Actions/Results를 분리해서 보여줍니다. Is_New=Yes는 이전 latest 파일 대비 신규 미션입니다."},
            {"Item": "Plan_Summary", "Description": "Action_Plan을 모드별로 압축한 요약입니다. 게임 시작 전 최종 확인용입니다."},
            {"Item": "Mode_Plan", "Description": "이전 방식의 모드별 압축표입니다. 참고용으로 유지합니다."},
            {"Item": "Seasonal_Mode_Plan", "Description": "Seasonal objective만 따로 모은 실행계획입니다. 장기 반복 미션을 별도로 볼 때 사용합니다."},
            {"Item": "Mode_Plan_All", "Description": "Seasonal까지 포함한 전체 실행계획입니다. 참고용입니다."},
            {"Item": "Starting_XI", "Description": "선발에 넣기만 하면 되는 조건입니다. 같은 속성은 최대 필요 수만 남깁니다."},
            {"Item": "Bench_or_Performer_Candidates", "Description": "직접 골/어시/기타 행동을 해야 할 가능성이 있는 선수 조건입니다. 선발 또는 교체로 준비하세요."},
            {"Item": "Goals / Assists", "Description": "누구로 골/어시를 해야 하는지 압축한 지시문입니다."},
            {"Item": "To do(기타)", "Description": "미션별 Win/Play/Complete/Clean Sheet도 누락하지 않고 표시합니다. 선발조건이 붙은 경우 ‘○○ 선수 n명 이상 선발하고 n승/플레이’로 연결해 보여줍니다."},
            {"Item": "Per_Match 의미", "Description": "'5경기 각 Goal 1회'처럼 표기되면 한 경기 몰아치기가 안 되고 별도 경기마다 필요합니다."},
            {"Item": "Conflict_Check", "Description": "모드 충돌, 같은 포지션 조건 충돌 등 자동 감지 가능한 주의사항입니다."},
            {"Item": "모드 묶기", "Description": "Squad Battles/Rivals/Champions/Live Events와 여기에 Rush가 추가된 문구는 실전 편의를 위해 같은 모드 그룹으로 묶습니다."},
            {"Item": "Other_Actions", "Description": "Clean Sheet 등 Goal/Assist가 아닌 특수 수행조건을 표시합니다. Outside Box Goal, Through Ball Assist처럼 골/어시 계열은 Goals/Assists로 분류합니다."},
            {"Item": "Parse_Check", "Description": "특수 키워드가 포함된 미션이 제대로 파싱됐는지 확인하는 시트입니다. CHECK가 있으면 Raw_Description 확인이 필요합니다."},
            {"Item": "신규미션 표시", "Description": "latest/objectives_latest.xlsx의 Mission_DB와 비교해 새로 추가된 미션의 To do 항목에는 (신규미션)을 붙입니다. 선발/교체 요건은 신규로 생겼거나 요구 수량이 증가한 경우에만 표시합니다."},
            {"Item": "기한 표시", "Description": "Collected_URLs의 카드 문구에서 Expires in 정보를 추출한 뒤, 실제 실행 지연과 무관하게 매일 02:00 KST를 기준으로 서울시간 마감시각을 계산합니다. 예: 마감 5/7 02:00"},
            {"Item": "주의", "Description": "규칙 기반 파싱이라 새 문장 유형은 일부 누락될 수 있습니다. Mission_DB의 Raw_Description을 함께 확인하세요."},
        ])
        guide.to_excel(writer, sheet_name="How_To_Read", index=False)

    format_workbook(out_path)


def format_workbook(path: Path):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    light_fill = PatternFill("solid", fgColor="F8FBFD")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9EAF7")

    widths = {
        "SB_Report": {"A": 24, "B": 60, "C": 58},
        "New_Report_Index": {"A": 18, "B": 46, "C": 18, "D": 12, "E": 18, "F": 48},
        "Combo_Report_Index": {"A": 18, "B": 20, "C": 46, "D": 16, "E": 12, "F": 58, "G": 48},
        "Mode_Report_Index": {"A": 18, "B": 46, "C": 18, "D": 12, "E": 18, "F": 48},
        "Combo_Report_Index": {"A": 18, "B": 20, "C": 46, "D": 16, "E": 12, "F": 58, "G": 48},
        "SB_Action_Detail": {
            "A": 12, "B": 34, "C": 20, "D": 30, "E": 26, "F": 28,
            "G": 28, "H": 34, "I": 24, "J": 12, "K": 12, "L": 16
        },
        "Action_Plan": {
            "A": 12, "B": 14, "C": 36, "D": 34, "E": 20, "F": 32, "G": 26,
            "H": 28, "I": 28, "J": 34, "K": 24, "L": 12, "M": 12, "N": 16,
            "O": 30, "P": 70, "Q": 48
        },
        "Seasonal_Action_Plan": {
            "A": 12, "B": 14, "C": 36, "D": 34, "E": 20, "F": 32, "G": 26,
            "H": 28, "I": 28, "J": 34, "K": 24, "L": 12, "M": 12, "N": 16,
            "O": 30, "P": 70, "Q": 48
        },
        "Action_Plan_All": {
            "A": 12, "B": 14, "C": 36, "D": 34, "E": 20, "F": 32, "G": 26,
            "H": 28, "I": 28, "J": 34, "K": 24, "L": 12, "M": 12, "N": 16,
            "O": 30, "P": 70, "Q": 48
        },
        "Plan_Summary": {
            "A": 36, "B": 12, "C": 12, "D": 34, "E": 28, "F": 34,
            "G": 34, "H": 34, "I": 28, "J": 42
        },
        "Mode_Plan": {
            "A": 36, "B": 12, "C": 12, "D": 34, "E": 28, "F": 34,
            "G": 34, "H": 28, "I": 28, "J": 38, "K": 40
        },
        "Seasonal_Mode_Plan": {
            "A": 36, "B": 12, "C": 12, "D": 34, "E": 28, "F": 34,
            "G": 34, "H": 28, "I": 28, "J": 38, "K": 40
        },
        "Mode_Plan_All": {
            "A": 36, "B": 12, "C": 12, "D": 34, "E": 28, "F": 34,
            "G": 34, "H": 28, "I": 28, "J": 38, "K": 40
        },
        "Mission_DB": {
            "A": 14, "B": 28, "C": 18, "D": 16, "E": 22, "F": 60, "G": 38, "H": 18,
            "I": 13, "J": 13, "K": 13, "L": 16, "M": 28, "N": 48, "O": 18, "P": 30
        },
        "Condition_DB": {
            "A": 14, "B": 18, "C": 20, "D": 20, "E": 24, "F": 12,
            "G": 12, "H": 13, "I": 42
        },
        "Conflict_Check": {"A": 26, "B": 32, "C": 28, "D": 38, "E": 48},
        "Parse_Check": {"A": 14, "B": 32, "C": 20, "D": 24, "E": 24, "F": 40, "G": 14, "H": 70},
        "Collected_URLs": {"A": 48, "B": 24, "C": 50, "D": 18},
        "How_To_Read": {"A": 24, "B": 90},
    }

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = True
        max_col = ws.max_column
        max_row = ws.max_row

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        if ws.title == "SB_Report" or ws.title.startswith("New_Report_") or ws.title.startswith("Mode_Report_") or ws.title.startswith("Combo_Report_"):
            for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="F8FBFD")
            for cell in ws["A"][1:]:
                cell.font = Font(bold=True, color="1F4E78")
                cell.fill = PatternFill("solid", fgColor="EAF3F8")
            if ws.max_column >= 3:
                for cell in ws["C"][1:]:
                    cell.font = Font(size=9, color="808080")
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
            ws.row_dimensions[1].height = 24

            if ((ws.title.startswith("New_Report_") and ws.title != "New_Report_Index")
                or (ws.title.startswith("Mode_Report_") and ws.title != "Mode_Report_Index")
                or (ws.title.startswith("Combo_Report_") and ws.title != "Combo_Report_Index")):
                ws.column_dimensions["A"].width = 24
                ws.column_dimensions["B"].width = 60
                ws.column_dimensions["C"].width = 58

        if ws.title in {"SB_Action_Detail", "Action_Plan", "Seasonal_Action_Plan", "Action_Plan_All"}:
            for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                for cell in row:
                    cell.fill = light_fill
            # Completed column
            for cell in ws["A"][1:]:
                cell.fill = PatternFill("solid", fgColor="E2F0D9")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # Is_New column, if present
            if ws.max_column >= 2 and ws.cell(row=1, column=2).value == "Is_New":
                for cell in ws["B"][1:]:
                    if str(cell.value).strip().lower() == "yes":
                        cell.fill = PatternFill("solid", fgColor="FFF2CC")
                        cell.font = Font(bold=True, color="C00000")
                    else:
                        cell.fill = PatternFill("solid", fgColor="F2F2F2")
                ws.column_dimensions["B"].width = 10

        if ws.title in {"Action_Plan", "Seasonal_Action_Plan", "Action_Plan_All"}:
            # Min matches column
            for cell in ws["L"][1:]:
                cell.fill = warn_fill
                cell.font = Font(bold=True)

        if ws.title == "SB_Action_Detail":
            # Min matches column is J in this compact sheet
            for cell in ws["J"][1:]:
                cell.fill = warn_fill
                cell.font = Font(bold=True)

        if ws.title in {"Plan_Summary", "Mode_Plan", "Seasonal_Mode_Plan", "Mode_Plan_All"}:
            for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                for cell in row:
                    cell.fill = light_fill
            # Highlight min matches
            for cell in ws["C"][1:]:
                cell.fill = warn_fill
                cell.font = Font(bold=True)

        for col, width in widths.get(ws.title, {}).items():
            ws.column_dimensions[col].width = width

        # Add autofilter
        ws.auto_filter.ref = ws.dimensions

    wb.save(path)


def update_analysis_workbook(out_path: Path, analysis_path: Path):
    # Optional: replace sheets in an existing workbook.
    # This is intentionally conservative: for now it simply saves the new workbook.
    # A later version can copy sheets into a permanent analysis workbook if needed.
    pass




def resolve_previous_objectives_file() -> Optional[Path]:
    """
    Previous-run comparison file.
    Priority:
      1) --previous "path/to/objectives_latest.xlsx"
      2) latest/objectives_latest.xlsx next to this script/repository root
    """
    prev_arg = get_cli_value("--previous")
    if prev_arg:
        p = Path(prev_arg).expanduser().resolve()
        return p if p.exists() else None

    default_prev = SCRIPT_DIR / "latest" / "objectives_latest.xlsx"
    return default_prev if default_prev.exists() else None


def mission_key_from_values(source_url, raw_description, mission_name="") -> str:
    source = canonical_url(str(source_url or ""))
    raw = clean_text(str(raw_description or "")).lower()
    name = clean_text(str(mission_name or "")).lower()
    return f"{source}||{raw}||{name}"


def mark_new_missions(df_mission: pd.DataFrame, previous_file: Optional[Path]) -> pd.DataFrame:
    """
    Add Is_New column by comparing current Mission_DB with previous Mission_DB.
    If there is no previous file, mark everything as No to avoid false 'new' labels.
    """
    if df_mission is None or df_mission.empty:
        return df_mission

    df = df_mission.copy()
    df["Is_New"] = "No"

    if previous_file is None or not previous_file.exists():
        print("Previous objectives file not found. Is_New will be No for all missions.")
        return df

    try:
        prev = pd.read_excel(previous_file, sheet_name="Mission_DB")
    except Exception as e:
        print(f"Previous objectives file could not be read: {previous_file} ({type(e).__name__}: {e})")
        return df

    prev_keys = set()
    for _, r in prev.iterrows():
        prev_keys.add(mission_key_from_values(
            r.get("Source_URL", ""),
            r.get("Raw_Description", ""),
            r.get("Mission_Name", ""),
        ))

    cur_keys = []
    for _, r in df.iterrows():
        cur_keys.append(mission_key_from_values(
            r.get("Source_URL", ""),
            r.get("Raw_Description", ""),
            r.get("Mission_Name", ""),
        ))

    df["Is_New"] = ["No" if k in prev_keys else "Yes" for k in cur_keys]
    print(f"New missions detected: {(df['Is_New'] == 'Yes').sum()}")
    return df



def clean_campaign_name(name: str) -> str:
    """
    Shorten FUT.GG campaign titles for report display.
    """
    s = clean_text(str(name or ""))
    s = re.sub(r"\s*-\s*EA SPORTS FC\s*\d+\s*Objectives\s*$", "", s, flags=re.I)
    s = re.sub(r"\s*-\s*EA SPORTS FC\s*\d+\s*$", "", s, flags=re.I)
    s = re.sub(r"\s*Objectives\s*$", "", s, flags=re.I)
    return clean_text(s)


def campaign_display_from_row(row) -> str:
    campaign = clean_campaign_name(str(row.get("Campaign", "")))
    expires = clean_text(str(row.get("Campaign_Expires", "")))
    deadline_kst = clean_text(str(row.get("Campaign_Deadline_KST", "")))
    display = campaign_expiry_display(expires, deadline_kst)
    if display and display.lower() != "nan":
        return f"{campaign} ({display})"
    return campaign if campaign else "-"


def report_lines_for_column(df: pd.DataFrame, col: str, marker: str = " (신규미션)") -> List[Dict]:
    """
    Return one report row per visible item so B열 내용 and C열 관련 캠페인이 1:1 대응됩니다.
    """
    rows = []
    if df is None or df.empty or col not in df.columns:
        return rows

    seen = set()
    for _, row in df.iterrows():
        is_new = str(row.get("Is_New", "No")).strip().lower() == "yes"
        campaign = campaign_display_from_row(row)
        for line in split_cell_lines(row.get(col, "")):
            text = line
            if is_new and col in {"Goals", "Assists", "Other_Actions"} and marker not in text:
                text = f"{text}{marker}"
            key = (text, campaign)
            if key not in seen:
                seen.add(key)
                rows.append({"내용": text, "관련 캠페인": campaign})
    return rows


def report_starting_rows(df: pd.DataFrame, marker: str = " (신규미션)") -> List[Dict]:
    """
    Starting_XI rows with conservative new/increased marker and 1:1 campaign mapping.
    If the same final requirement is supported by multiple campaigns, one row per
    campaign is kept so the source remains clear.
    """
    if df is None or df.empty or "Starting_XI" not in df.columns:
        return []

    old_max = {}
    all_max = {}
    source_rows = {}

    for idx, row in df.iterrows():
        is_new = str(row.get("Is_New", "No")).strip().lower() == "yes"
        for line in split_cell_lines(row.get("Starting_XI", "")):
            parsed = parse_ge_requirement_line(line)
            if not parsed:
                key = ("RAW", line)
                label = line
                cnt = 1
            else:
                key, label, cnt = parsed

            all_max[key] = max(all_max.get(key, 0), cnt)
            source_rows.setdefault(key, []).append((idx, row, cnt, label))
            if not is_new:
                old_max[key] = max(old_max.get(key, 0), cnt)

    out = []
    seen = set()
    for key in sorted(all_max.keys(), key=lambda x: str(x)):
        max_cnt = all_max[key]
        is_new_or_increased = max_cnt > old_max.get(key, 0)

        # show campaigns that contribute to the max requirement. If none, fallback all.
        candidates = [x for x in source_rows.get(key, []) if x[2] == max_cnt]
        if not candidates:
            candidates = source_rows.get(key, [])

        for _, row, _, label in candidates:
            if key[0] == "RAW":
                text = label
            else:
                text = f"{label} ≥{max_cnt}"
            if is_new_or_increased and marker not in text:
                text += marker
            campaign = campaign_display_from_row(row)
            dedupe = (text, campaign)
            if dedupe not in seen:
                seen.add(dedupe)
                out.append({"내용": text, "관련 캠페인": campaign})
    return out


def report_performer_candidate_rows(df: pd.DataFrame, marker: str = " (신규미션)") -> List[Dict]:
    """
    Performer candidate rows excluding values already required in Starting_XI,
    with conservative new marker and 1:1 campaign mapping.
    """
    if df is None or df.empty:
        return []

    starting_text = unique_join_starting_with_new(df, marker="")
    starting_values = get_starting_requirement_values(starting_text)

    old_vals = set()
    if "Performer_Candidates" in df.columns:
        for _, row in df.iterrows():
            is_new = str(row.get("Is_New", "No")).strip().lower() == "yes"
            vals = [v for v in split_cell_lines(row.get("Performer_Candidates", "")) if v not in starting_values]
            if not is_new:
                old_vals.update(vals)

    out = []
    seen = set()
    if "Performer_Candidates" in df.columns:
        for _, row in df.iterrows():
            is_new = str(row.get("Is_New", "No")).strip().lower() == "yes"
            campaign = campaign_display_from_row(row)
            for v in split_cell_lines(row.get("Performer_Candidates", "")):
                if v in starting_values:
                    continue
                text = v
                if is_new and v not in old_vals and marker not in text:
                    text += marker
                key = (text, campaign)
                if key not in seen:
                    seen.add(key)
                    out.append({"내용": text, "관련 캠페인": campaign})
    return out


def append_report_section(rows: List[Dict], section: str, items: List[Dict], empty_text: str = "-"):
    """
    Append report rows in one-item-per-row form.
    First row carries the section title, following rows leave 구분 blank.
    """
    if not items:
        rows.append({"구분": section, "내용": empty_text, "관련 캠페인": "-"})
        return

    for i, item in enumerate(items):
        rows.append({
            "구분": section if i == 0 else "",
            "내용": item.get("내용", "-"),
            "관련 캠페인": item.get("관련 캠페인", "-"),
        })


def unique_join_from_column_with_new(df: pd.DataFrame, col: str, marker: str = " (신규미션)") -> str:
    """
    Same as unique_join_from_column, but lines originating from Is_New=Yes rows
    receive a marker such as ' (신규미션)'.
    """
    if df is None or df.empty or col not in df.columns:
        return "-"

    old_lines = []
    new_lines = []

    for _, row in df.iterrows():
        vals = split_cell_lines(row.get(col, ""))
        if not vals:
            continue
        if str(row.get("Is_New", "No")).strip().lower() == "yes":
            new_lines.extend(vals)
        else:
            old_lines.extend(vals)

    new_set = set(unique_list(new_lines))
    all_lines = unique_list(old_lines + new_lines)

    out = []
    for line in all_lines:
        if line in new_set and marker not in line:
            out.append(f"{line}{marker}")
        else:
            out.append(line)

    return "\n".join(out) if out else "-"


def load_config() -> Dict:
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
        except Exception as e:
            print(f"Config file read failed: {CONFIG_FILE} ({type(e).__name__}: {e})")
    return {}


def get_cli_value(flag: str) -> Optional[str]:
    if flag not in sys.argv:
        return None
    idx = sys.argv.index(flag)
    if idx + 1 >= len(sys.argv):
        return None
    return sys.argv[idx + 1]


def resolve_out_dir() -> Path:
    """
    Output folder priority:
      1) --out "folder path"
      2) futgg_objective_config.json: {"out_dir": "..."}
      3) same folder as this script
    """
    cfg = load_config()

    out_arg = get_cli_value("--out")
    if out_arg:
        return Path(out_arg).expanduser().resolve()

    cfg_out = cfg.get("out_dir") if isinstance(cfg, dict) else None
    if cfg_out:
        return Path(str(cfg_out)).expanduser().resolve()

    return DEFAULT_OUT_DIR


def resolve_headless(default_headless: bool = True) -> bool:
    """
    Browser visibility priority:
      1) --show  => headless False
      2) --headless => headless True
      3) config headless true/false
      4) default
    """
    if "--show" in sys.argv:
        return False
    if "--headless" in sys.argv:
        return True

    cfg = load_config()
    if isinstance(cfg, dict) and "headless" in cfg:
        return bool(cfg["headless"])

    return default_headless


def main():
    today = datetime.now().strftime("%Y%m%d")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    out_dir = resolve_out_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"futgg_objectives_plan_{timestamp}.xlsx"

    headless = resolve_headless(default_headless=True)

    print("FUT.GG Objective crawler started.")
    print(f"Script folder: {SCRIPT_DIR}")
    print(f"Output folder: {out_dir}")
    print(f"Output file will be: {out_path}")
    print(f"Collection time KST: {datetime.now(KST).strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Deadline calculation base KST: {scheduled_deadline_base_kst().strftime('%Y-%m-%d %H:%M:%S')}")
    print("Collecting objective URLs...")
    links = collect_objective_links(headless=headless)

    all_raw_missions = []
    errors = []
    raw_pages = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(viewport={"width": 1450, "height": 1200})

        for idx, link in enumerate(links, start=1):
            print(f"[{idx}/{len(links)}] {link['URL']}")
            try:
                mission_rows, raw_short = scrape_objective_page(context, link, idx)
                raw_pages.append({"URL": link["URL"], "Raw_Page_Text_Short": raw_short})
                if not mission_rows:
                    errors.append({"URL": link["URL"], "Error": "No missions extracted or expired page skipped."})
                all_raw_missions.extend(mission_rows)
            except Exception as e:
                errors.append({"URL": link["URL"], "Error": f"{type(e).__name__}: {e}"})

        browser.close()

    mission_db_rows = []
    condition_rows = []

    for m in all_raw_missions:
        mission_row, conds = parse_mission_conditions(m)
        mission_db_rows.append(mission_row)
        condition_rows.extend(conds)

    df_mission = pd.DataFrame(mission_db_rows)
    df_conditions = pd.DataFrame(condition_rows)

    previous_file = resolve_previous_objectives_file()
    if previous_file:
        print(f"Previous objectives file for new-mission check: {previous_file}")
    df_mission = mark_new_missions(df_mission, previous_file)

    # 실행계획은 기본적으로 Seasonal을 제외한 버전과 Seasonal만 따로 보는 버전으로 나눕니다.
    if EXCLUDE_SEASONAL_FROM_MAIN_PLAN and not df_mission.empty:
        non_seasonal_ids = set(df_mission.loc[df_mission["Objective_Type"] != "Seasonal", "Mission_ID"])
        seasonal_ids = set(df_mission.loc[df_mission["Objective_Type"] == "Seasonal", "Mission_ID"])

        df_mission_main = df_mission[df_mission["Mission_ID"].isin(non_seasonal_ids)].copy()
        df_conditions_main = df_conditions[df_conditions["Mission_ID"].isin(non_seasonal_ids)].copy()

        df_mission_seasonal = df_mission[df_mission["Mission_ID"].isin(seasonal_ids)].copy()
        df_conditions_seasonal = df_conditions[df_conditions["Mission_ID"].isin(seasonal_ids)].copy()
    else:
        df_mission_main = df_mission
        df_conditions_main = df_conditions
        df_mission_seasonal = pd.DataFrame()
        df_conditions_seasonal = pd.DataFrame()

    df_action = build_action_plan(df_mission_main, df_conditions_main)
    df_action_seasonal = build_action_plan(df_mission_seasonal, df_conditions_seasonal)
    df_action_all = build_action_plan(df_mission, df_conditions)

    df_plan = build_plan_summary_from_action_plan(df_action)
    df_seasonal_plan = build_plan_summary_from_action_plan(df_action_seasonal)
    df_plan_all = build_plan_summary_from_action_plan(df_action_all)

    df_conflict = build_conflict_check(df_plan, df_conditions_main)
    df_links = pd.DataFrame(links)

    write_excel(out_path, df_mission, df_conditions, df_plan, df_conflict, df_links, errors,
                df_seasonal_plan=df_seasonal_plan, df_plan_all=df_plan_all,
                df_action=df_action, df_action_seasonal=df_action_seasonal, df_action_all=df_action_all)

    print()
    print(f"Done: {out_path}")
    print(f"Objective pages: {len(links)}")
    print(f"Missions parsed: {len(df_mission)}")
    print(f"Conditions parsed: {len(df_conditions)}")
    print("Mode_Plan의 Covered_Missions에는 캠페인명이 중복 없이 표시됩니다.")
    print("Squad Battles/Rivals/Champions/Live Events 계열은 Rush 포함 여부와 관계없이 같은 그룹으로 묶습니다.")
    if not df_plan.empty:
        print("Mode plans excluding Seasonal:")
        print(df_plan[["Mode_Group", "Min_Matches"]].to_string(index=False))
    if 'df_seasonal_plan' in locals() and not df_seasonal_plan.empty:
        print("Seasonal mode plans saved separately:")
        print(df_seasonal_plan[["Mode_Group", "Min_Matches"]].to_string(index=False))


if __name__ == "__main__":
    main()
