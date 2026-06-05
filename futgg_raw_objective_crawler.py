# futgg_raw_objective_crawler.py
# FUT.GG Objectives RAW crawler
#
# Purpose
# - Crawl active FUT.GG objective pages on a schedule.
# - Preserve objective wording as close to FUT.GG original as possible.
# - Do NOT classify Goals/Assists/Other, do NOT build posters, do NOT generate parsed mission guides.
# - Generate upload-friendly RAW xlsx/json files for ChatGPT/manual interpretation.
#
# GitHub Actions example:
#   python -u futgg_raw_objective_crawler.py --out outputs --headless
#
# Requirements:
#   pandas openpyxl beautifulsoup4 playwright
#   python -m playwright install chromium

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from collections import defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

START_URL = "https://www.fut.gg/objectives/"
BASE_URL = "https://www.fut.gg"
KST = timezone(timedelta(hours=9))
DEADLINE_BASE_HOUR_KST = 2
DEADLINE_BASE_MINUTE_KST = 0

OBJECTIVE_PATHS = (
    "campaigns",
    "live-events",
    "seasonal",
    "milestones",
    "fc-pro",
)

NOISE_LINES = {
    "home", "objectives", "players", "sbc", "squads", "evolutions", "login", "sign up",
    "rewards", "requirements", "details", "objectives", "playstyles", "roles",
}

ACTION_START_RE = re.compile(
    r"^(Win|Play|Score|Assist|Complete|Keep|Earn|Make|Get|Perform|Record|Concede|Claim|Finish|Achieve|Watch|Tune|View|Link|Submit|Build)\b",
    re.I,
)

REWARD_HINT_RE = re.compile(
    r"\b(Pack|Player Pick|Pick|Player|SP|XP|Coins?|Coin Boost|Evo Unlock|Evolution|Token|Loan|Kit|Badge|Tifo|FoF|EVO|Rare Gold|TOTS|TOTW|UT Champions|Draft Token|Consumable)\b",
    re.I,
)

MODE_HINT_RE = re.compile(
    r"\b(Squad Battles|Rivals|Champions|Live Events?|Rush|Ultimate Team|FC Pro|Twitch|YouTube|broadcast|stream|Semi-Pro|World Class|Professional|Legendary|difficulty)\b",
    re.I,
)


def clean_text(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def canonical_url(url: str) -> str:
    if not url:
        return ""
    if url.startswith("/"):
        url = BASE_URL + url
    return url.split("?", 1)[0].rstrip("/") + "/"


def is_objective_detail_url(url: str) -> bool:
    u = canonical_url(url)
    paths = "|".join(map(re.escape, OBJECTIVE_PATHS))
    return re.search(rf"/objectives/({paths})/[0-9]+-[^/]+/$", u) is not None


def slug_title(url: str) -> str:
    slug = canonical_url(url).rstrip("/").split("/")[-1]
    slug = re.sub(r"^\d+-", "", slug)
    return " ".join(w.capitalize() for w in slug.split("-"))


def infer_objective_type(url: str) -> str:
    u = canonical_url(url)
    if "/campaigns/" in u:
        return "Campaign"
    if "/live-events/" in u:
        return "Live Events"
    if "/seasonal/" in u:
        return "Seasonal"
    if "/milestones/" in u:
        return "Milestone"
    if "/fc-pro/" in u:
        return "FC Pro"
    return "Objective"


def is_expired_text(text: str) -> bool:
    # FUT.GG active cards usually say "Expires in ..."; inactive cards can say "Expired".
    return re.search(r"\bExpired\b", clean_text(text), re.I) is not None


def parse_expires_text(text: str) -> str:
    t = clean_text(text)
    if not t:
        return ""
    m = re.search(
        r"\bExpires\s+in\s+((?:[0-9]+\s*(?:days?|hours?|hrs?|minutes?|mins?)\s*){1,4})",
        t,
        re.I,
    )
    if m:
        return "Expires in " + clean_text(m.group(1))
    m = re.search(r"\bExpires\s+in\s+(.+?)(?=\s+[0-9,]+\s+total|\s+[0-9]+\s+Objectives|\s+[0-9]+%|$)", t, re.I)
    if m:
        return "Expires in " + clean_text(m.group(1))
    return ""


def scheduled_deadline_base_kst(now_dt: Optional[datetime] = None) -> datetime:
    if now_dt is None:
        now_dt = datetime.now(KST)
    elif now_dt.tzinfo is None:
        now_dt = now_dt.replace(tzinfo=KST)
    else:
        now_dt = now_dt.astimezone(KST)

    base = now_dt.replace(
        hour=DEADLINE_BASE_HOUR_KST,
        minute=DEADLINE_BASE_MINUTE_KST,
        second=0,
        microsecond=0,
    )
    if now_dt < base:
        base -= timedelta(days=1)
    return base


def expiry_deadline_kst(expires_text: str, base_dt: Optional[datetime] = None) -> Tuple[str, str]:
    text = clean_text(expires_text)
    if not text:
        return "", ""
    if base_dt is None:
        base_dt = scheduled_deadline_base_kst()
    elif base_dt.tzinfo is None:
        base_dt = base_dt.replace(tzinfo=KST)
    else:
        base_dt = base_dt.astimezone(KST)

    duration = timedelta(0)
    for num, unit in re.findall(r"([0-9]+)\s*(days?|hours?|hrs?|minutes?|mins?)", text, re.I):
        n = int(num)
        u = unit.lower()
        if u.startswith("day"):
            duration += timedelta(days=n)
        elif u.startswith("hour") or u.startswith("hr"):
            duration += timedelta(hours=n)
        elif u.startswith("minute") or u.startswith("min"):
            duration += timedelta(minutes=n)

    if duration == timedelta(0):
        return "", ""

    deadline = base_dt + duration
    display = deadline.strftime("%-m/%-d %H:%M") if sys.platform != "win32" else deadline.strftime("%#m/%#d %H:%M")
    return display, deadline.isoformat(timespec="minutes")


def split_lines(text: str) -> List[str]:
    lines: List[str] = []
    for raw in (text or "").replace("\r", "\n").split("\n"):
        line = clean_text(raw)
        if not line:
            continue
        if line.lower() in NOISE_LINES:
            continue
        lines.append(line)
    return lines


def looks_like_description(line: str) -> bool:
    t = clean_text(line)
    if not t:
        return False
    if not ACTION_START_RE.search(t):
        return False
    # Avoid short mission-card headings like "Win 2" or "Score in 2" unless they contain context.
    if re.fullmatch(r"(Win|Play|Score|Assist|Keep|Clean Sheet|Complete)\s+(?:in\s+)?[0-9]+", t, re.I):
        return False
    return bool(re.search(
        r"\b(matches?|goals?|assists?|clean sheets?|Squad Battles|Rivals|Champions|Live Events?|Rush|starting 11|using|with|while having|difficulty|broadcast|stream|FC Pro|Twitch|YouTube|watch|Ultimate Team)\b",
        t,
        re.I,
    ))


def looks_like_reward(line: str) -> bool:
    t = clean_text(line)
    if not t:
        return False
    if looks_like_description(t):
        return False
    return bool(REWARD_HINT_RE.search(t))


def looks_like_mission_name(line: str) -> bool:
    t = clean_text(line)
    if not t or len(t) > 80:
        return False
    if looks_like_reward(t):
        return False
    if looks_like_description(t):
        return False
    if re.search(r"^(Win|Play|Score|Assist|Keep|Clean Sheet|Complete|Earn|Get)\b", t, re.I):
        return True
    # Custom short names like "Curl it in", "In to the Box", "Stellar Defense".
    return bool(re.search(r"[A-Za-z]", t)) and not re.search(r"https?://", t)


def split_description_and_reward(line: str) -> Tuple[str, str]:
    """
    FUT.GG sometimes combines objective description and reward with slash separators.
    Keep the description original and move reward-looking suffixes into Reward_Text_Raw.
    """
    t = clean_text(line)
    if " / " not in t:
        return t, ""
    parts = [clean_text(p) for p in t.split(" / ") if clean_text(p)]
    if not parts:
        return t, ""
    if ACTION_START_RE.search(parts[0]):
        reward_parts = [p for p in parts[1:] if looks_like_reward(p)]
        return parts[0], " / ".join(reward_parts)
    return t, ""


def nearby_reward_lines(lines: List[str], desc_idx: int) -> str:
    found: List[str] = []
    # Reward is usually before or after the full description depending on page layout.
    scan_indices = list(range(max(0, desc_idx - 4), desc_idx)) + list(range(desc_idx + 1, min(len(lines), desc_idx + 7)))
    for j in scan_indices:
        cand = clean_text(lines[j])
        if not cand:
            continue
        desc_part, reward_part = split_description_and_reward(cand)
        if reward_part and desc_part != cand:
            continue
        if looks_like_reward(cand):
            found.append(cand)
    return " / ".join(unique_list(found)[:4])


def extract_mode_text(desc: str) -> str:
    d = clean_text(desc)
    modes: List[str] = []
    if re.search(r"Squad Battles", d, re.I):
        modes.append("Squad Battles")
    if re.search(r"Rivals", d, re.I):
        modes.append("Rivals")
    if re.search(r"Champions", d, re.I):
        modes.append("Champions")
    if re.search(r"Live Events?", d, re.I):
        modes.append("Live Events")
    if re.search(r"\bRush\b", d, re.I):
        modes.append("Rush")
    if re.search(r"any Ultimate Team", d, re.I):
        modes.append("Any Ultimate Team mode")
    if re.search(r"FC Pro|Twitch|YouTube|broadcast|stream|Watch|Tune", d, re.I):
        modes.append("Watch / FC Pro")
    ev = re.search(r"\bin the (.+? event)\b", d, re.I)
    if ev:
        modes.append(clean_text(ev.group(1)))
    return " / ".join(unique_list(modes))


def extract_difficulty_text(desc: str) -> str:
    d = clean_text(desc)
    m = re.search(r"Min\.?\s*([A-Za-z -]+?)\s+difficulty", d, re.I)
    if m:
        return "Min. " + clean_text(m.group(1))
    m = re.search(r"on\s+min\.?\s*([A-Za-z -]+?)\s+difficulty", d, re.I)
    if m:
        return "Min. " + clean_text(m.group(1))
    return ""


def unique_list(items: List[str]) -> List[str]:
    seen = set()
    out = []
    for x in items:
        x = clean_text(x)
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


def objective_key(url: str, objective_text: str, mission_name: str = "") -> str:
    return f"{canonical_url(url)}||{clean_text(objective_text).lower()}||{clean_text(mission_name).lower()}"


def collect_objective_links(headless: bool = True, max_click_rounds: int = 8) -> List[Dict[str, Any]]:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        page = browser.new_page(viewport={"width": 1450, "height": 1200})
        page.goto(START_URL, wait_until="networkidle", timeout=60000)
        time.sleep(2)

        # Expand visible lists if FUT.GG hides cards behind buttons.
        for _ in range(max_click_rounds):
            clicked = False
            for pat in [r"Show all", r"View all", r"Load more", r"Show more"]:
                try:
                    loc = page.get_by_text(re.compile(pat, re.I))
                    for i in range(min(loc.count(), 8)):
                        try:
                            item = loc.nth(i)
                            if item.is_visible(timeout=500):
                                item.click(timeout=1500, force=True)
                                clicked = True
                                time.sleep(0.4)
                        except Exception:
                            pass
                except Exception:
                    pass
            if not clicked:
                break
            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass

        raw = page.eval_on_selector_all(
            "a[href]",
            """
            els => els.map(a => {
                let parent = a;
                for (let i = 0; i < 6; i++) {
                    if (!parent.parentElement) break;
                    parent = parent.parentElement;
                    const txt = (parent.innerText || parent.textContent || '').trim();
                    if (txt.length > 40) break;
                }
                return {
                    href: a.href,
                    link_text: (a.innerText || a.textContent || '').trim(),
                    parent_text: (parent.innerText || parent.textContent || '').trim()
                };
            })
            """,
        )
        browser.close()

    rows: List[Dict[str, Any]] = []
    seen = set()
    for item in raw:
        url = canonical_url(item.get("href", ""))
        if not is_objective_detail_url(url):
            continue
        link_text = clean_text(item.get("link_text", ""))
        parent_text = clean_text(item.get("parent_text", ""))
        if is_expired_text(parent_text + " " + link_text):
            continue
        expires = parse_expires_text(parent_text + " " + link_text)
        deadline_kst, deadline_iso = expiry_deadline_kst(expires)
        if url not in seen:
            seen.add(url)
            rows.append({
                "URL": url,
                "Objective_Type": infer_objective_type(url),
                "Link_Text": link_text,
                "Parent_Text": parent_text,
                "Campaign_Expires": expires,
                "Campaign_Deadline_KST": deadline_kst,
                "Campaign_Deadline_ISO": deadline_iso,
            })
    return rows


def extract_objectives_from_body(body: str) -> List[Dict[str, Any]]:
    lines = split_lines(body)
    objectives: List[Dict[str, Any]] = []
    seen_desc = set()

    for i, line in enumerate(lines):
        desc, reward_suffix = split_description_and_reward(line)
        if not looks_like_description(desc):
            continue
        if desc in seen_desc:
            continue
        seen_desc.add(desc)

        # Nearest mission card title above the description.
        name = ""
        for j in range(i - 1, max(-1, i - 7), -1):
            cand = clean_text(lines[j])
            if looks_like_mission_name(cand):
                name = cand
                break
        if not name:
            name = desc[:60]

        reward_near = nearby_reward_lines(lines, i)
        reward = " / ".join([x for x in [reward_suffix, reward_near] if x])

        objectives.append({
            "Objective_Order": len(objectives) + 1,
            "Objective_Name_Raw": name,
            "Objective_Text_Raw": desc,
            "Reward_Text_Raw": reward,
            "Mode_Text_Raw": extract_mode_text(desc),
            "Difficulty_Text_Raw": extract_difficulty_text(desc),
            "Extraction_Method": "description_line_with_backward_title",
        })

    # Fallback: preserve action-looking lines even if not enough context is detected.
    if not objectives:
        seen = set()
        for i, line in enumerate(lines):
            if ACTION_START_RE.search(line) and not looks_like_reward(line):
                text = clean_text(line)
                if text in seen:
                    continue
                seen.add(text)
                objectives.append({
                    "Objective_Order": len(objectives) + 1,
                    "Objective_Name_Raw": text[:60],
                    "Objective_Text_Raw": text,
                    "Reward_Text_Raw": nearby_reward_lines(lines, i),
                    "Mode_Text_Raw": extract_mode_text(text),
                    "Difficulty_Text_Raw": extract_difficulty_text(text),
                    "Extraction_Method": "action_line_fallback",
                })

    return objectives


def scrape_objective_page(context, link_row: Dict[str, Any], seq: int) -> Tuple[Dict[str, Any], List[Dict[str, Any]], str]:
    url = link_row["URL"]
    page = context.new_page()
    try:
        page.goto(url, wait_until="networkidle", timeout=60000)
        time.sleep(1)
        body = page.inner_text("body", timeout=15000)

        if is_expired_text(body):
            raise RuntimeError("Expired page skipped")

        title = ""
        try:
            h1s = page.locator("h1").all_inner_texts()
            if h1s:
                title = clean_text(h1s[0])
        except Exception:
            pass
        if not title:
            title = slug_title(url)

        expires = link_row.get("Campaign_Expires", "") or parse_expires_text(body)
        deadline_kst = link_row.get("Campaign_Deadline_KST", "")
        deadline_iso = link_row.get("Campaign_Deadline_ISO", "")
        if expires and not deadline_kst:
            deadline_kst, deadline_iso = expiry_deadline_kst(expires)

        group = {
            "Group_ID": f"G{seq:03d}",
            "Mission_Group_Name": title,
            "Objective_Type": link_row.get("Objective_Type", infer_objective_type(url)),
            "URL": url,
            "Campaign_Expires": expires,
            "Campaign_Deadline_KST": deadline_kst,
            "Campaign_Deadline_ISO": deadline_iso,
            "Page_Text_Raw_Short": clean_text(body)[:5000],
        }

        objectives = extract_objectives_from_body(body)
        for obj in objectives:
            obj.update({
                "Group_ID": group["Group_ID"],
                "Mission_Group_Name": title,
                "Objective_Type": group["Objective_Type"],
                "URL": url,
                "Campaign_Expires": expires,
                "Campaign_Deadline_KST": deadline_kst,
                "Campaign_Deadline_ISO": deadline_iso,
            })
            obj["Objective_Key"] = objective_key(url, obj.get("Objective_Text_Raw", ""), obj.get("Objective_Name_Raw", ""))
        return group, objectives, clean_text(body)[:12000]
    finally:
        page.close()


def find_previous_file(out_dir: Path, explicit_previous: Optional[str] = None) -> Optional[Path]:
    if explicit_previous:
        p = Path(explicit_previous).expanduser().resolve()
        return p if p.exists() else None

    # Useful when GitHub checkout has previous latest files.
    candidates = [
        Path.cwd() / "latest" / "objectives_raw_latest.xlsx",
        Path.cwd() / "latest" / "objectives_latest.xlsx",
        out_dir / "objectives_raw_latest.xlsx",
        out_dir / "objectives_latest.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def previous_keys_from_file(previous_file: Optional[Path]) -> set:
    if not previous_file or not previous_file.exists():
        return set()
    try:
        # Prefer the raw sheet from this script, fallback to older Mission_DB sheet.
        xls = pd.ExcelFile(previous_file)
        if "Objective_Raw" in xls.sheet_names:
            df_prev = pd.read_excel(previous_file, sheet_name="Objective_Raw")
            if "Objective_Key" in df_prev.columns:
                return set(df_prev["Objective_Key"].dropna().astype(str))
            return set(
                objective_key(r.get("URL", ""), r.get("Objective_Text_Raw", ""), r.get("Objective_Name_Raw", ""))
                for _, r in df_prev.iterrows()
            )
        if "Mission_DB" in xls.sheet_names:
            df_prev = pd.read_excel(previous_file, sheet_name="Mission_DB")
            return set(
                objective_key(r.get("Source_URL", r.get("URL", "")), r.get("Raw_Description", ""), r.get("Mission_Name", ""))
                for _, r in df_prev.iterrows()
            )
    except Exception as exc:
        print(f"Previous file read failed: {previous_file} ({type(exc).__name__}: {exc})")
    return set()


def apply_new_flags(df_groups: pd.DataFrame, df_objectives: pd.DataFrame, prev_keys: set) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df_o = df_objectives.copy()
    if df_o.empty:
        df_o["Is_New"] = []
    elif not prev_keys:
        # First run safety: avoid marking everything new.
        df_o["Is_New"] = "No"
    else:
        df_o["Is_New"] = df_o["Objective_Key"].astype(str).map(lambda k: "No" if k in prev_keys else "Yes")

    new_group_ids = set(df_o.loc[df_o.get("Is_New", "No") == "Yes", "Group_ID"].astype(str)) if not df_o.empty else set()
    df_g = df_groups.copy()
    if df_g.empty:
        df_g["Is_New_Group"] = []
        df_g["New_Objective_Count"] = []
    else:
        df_g["Is_New_Group"] = df_g["Group_ID"].astype(str).map(lambda gid: "Yes" if gid in new_group_ids else "No")
        counts = df_o[df_o.get("Is_New", "No") == "Yes"].groupby("Group_ID").size().to_dict() if not df_o.empty else {}
        df_g["New_Objective_Count"] = df_g["Group_ID"].astype(str).map(lambda gid: int(counts.get(gid, 0)))
    return df_g, df_o


def build_manifest(df_groups: pd.DataFrame, df_objectives: pd.DataFrame, errors: List[Dict[str, Any]]) -> pd.DataFrame:
    return pd.DataFrame([
        {"Item": "Generated_At_KST", "Value": datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")},
        {"Item": "Deadline_Base_KST", "Value": scheduled_deadline_base_kst().strftime("%Y-%m-%d %H:%M:%S")},
        {"Item": "Mission_Group_Count", "Value": len(df_groups)},
        {"Item": "Objective_Row_Count", "Value": len(df_objectives)},
        {"Item": "New_Objective_Count", "Value": int((df_objectives.get("Is_New", pd.Series(dtype=str)) == "Yes").sum()) if not df_objectives.empty else 0},
        {"Item": "Error_Count", "Value": len(errors)},
        {"Item": "Purpose", "Value": "RAW collection only. Do not trust this file as parsed mission guide; upload it to ChatGPT/manual review for route optimization."},
    ])


def to_json_payload(df_manifest: pd.DataFrame, df_groups: pd.DataFrame, df_objectives: pd.DataFrame, df_links: pd.DataFrame, errors: List[Dict[str, Any]]) -> Dict[str, Any]:
    groups = []
    obj_by_gid = defaultdict(list)
    for _, r in df_objectives.iterrows():
        obj_by_gid[str(r.get("Group_ID", ""))].append({
            "order": int(r.get("Objective_Order") or 0),
            "name_raw": clean_text(r.get("Objective_Name_Raw", "")),
            "text_raw": clean_text(r.get("Objective_Text_Raw", "")),
            "reward_raw": clean_text(r.get("Reward_Text_Raw", "")),
            "mode_raw": clean_text(r.get("Mode_Text_Raw", "")),
            "difficulty_raw": clean_text(r.get("Difficulty_Text_Raw", "")),
            "is_new": clean_text(r.get("Is_New", "")),
            "objective_key": clean_text(r.get("Objective_Key", "")),
        })

    for _, g in df_groups.iterrows():
        gid = str(g.get("Group_ID", ""))
        groups.append({
            "group_id": gid,
            "mission_group_name": clean_text(g.get("Mission_Group_Name", "")),
            "objective_type": clean_text(g.get("Objective_Type", "")),
            "url": clean_text(g.get("URL", "")),
            "campaign_expires": clean_text(g.get("Campaign_Expires", "")),
            "campaign_deadline_kst": clean_text(g.get("Campaign_Deadline_KST", "")),
            "campaign_deadline_iso": clean_text(g.get("Campaign_Deadline_ISO", "")),
            "is_new_group": clean_text(g.get("Is_New_Group", "")),
            "new_objective_count": int(g.get("New_Objective_Count") or 0),
            "objectives": sorted(obj_by_gid.get(gid, []), key=lambda x: x["order"]),
        })

    return {
        "manifest": {str(r["Item"]): r["Value"] for _, r in df_manifest.iterrows()},
        "missions": groups,
        "collected_urls": df_links.to_dict(orient="records"),
        "errors": errors,
    }


def write_outputs(out_dir: Path, df_manifest: pd.DataFrame, df_groups: pd.DataFrame, df_objectives: pd.DataFrame, df_links: pd.DataFrame, df_raw_pages: pd.DataFrame, errors: List[Dict[str, Any]], timestamp: str) -> Tuple[Path, Path]:
    xlsx_path = out_dir / f"futgg_objectives_raw_{timestamp}.xlsx"
    json_path = out_dir / f"futgg_objectives_raw_{timestamp}.json"

    df_errors = pd.DataFrame(errors)
    if df_errors.empty:
        df_errors = pd.DataFrame(columns=["URL", "Error"])

    guide = pd.DataFrame([
        {"Item": "핵심 시트", "Description": "Objective_Raw: 세부목표 1행=1원문. 포스터/루트 정리는 이 시트를 업로드해서 해석한다."},
        {"Item": "Mission_Groups", "Description": "미션 묶음 단위. URL, 유형, 마감, 신규 세부목표 수를 확인한다."},
        {"Item": "원칙", "Description": "이 파일은 원문 수집용이다. Goals/Assists/Other 자동 분류를 하지 않는다."},
        {"Item": "Is_New", "Description": "이전 latest raw/objectives 파일과 Objective_Key를 비교한다. 이전 파일이 없으면 전체 No 처리한다."},
        {"Item": "Mode_Text_Raw", "Description": "원문에서 감지한 모드 단서이다. 최종 루트 판단은 원문 전체를 보고 한다."},
        {"Item": "주의", "Description": "FUT.GG 페이지 구조가 바뀌면 Extract_Method=action_line_fallback 또는 Errors를 확인한다."},
    ])

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_manifest.to_excel(writer, sheet_name="Manifest", index=False)
        df_groups.to_excel(writer, sheet_name="Mission_Groups", index=False)
        df_objectives.to_excel(writer, sheet_name="Objective_Raw", index=False)
        df_links.to_excel(writer, sheet_name="Collected_URLs", index=False)
        df_raw_pages.to_excel(writer, sheet_name="Raw_Page_Text", index=False)
        df_errors.to_excel(writer, sheet_name="Errors", index=False)
        guide.to_excel(writer, sheet_name="How_To_Read", index=False)

    format_workbook(xlsx_path)

    payload = to_json_payload(df_manifest, df_groups, df_objectives, df_links, errors)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    # Also write stable latest-named copies into the output folder for workflow convenience.
    latest_xlsx = out_dir / "objectives_raw_latest.xlsx"
    latest_json = out_dir / "objectives_raw_latest.json"
    latest_xlsx.write_bytes(xlsx_path.read_bytes())
    latest_json.write_text(json_path.read_text(encoding="utf-8"), encoding="utf-8")

    return xlsx_path, json_path


def format_workbook(path: Path):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="102A56")
    header_font = Font(color="FFFFFF", bold=True)
    light_fill = PatternFill("solid", fgColor="F7FBFF")
    new_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9EAF7")

    widths = {
        "Manifest": {"A": 26, "B": 90},
        "Mission_Groups": {"A": 12, "B": 42, "C": 16, "D": 58, "E": 20, "F": 18, "G": 26, "H": 12, "I": 18, "J": 90},
        "Objective_Raw": {"A": 12, "B": 16, "C": 36, "D": 16, "E": 58, "F": 12, "G": 30, "H": 86, "I": 46, "J": 38, "K": 18, "L": 20, "M": 18, "N": 26, "O": 58, "P": 26},
        "Collected_URLs": {"A": 60, "B": 16, "C": 30, "D": 70, "E": 20, "F": 18, "G": 26},
        "Raw_Page_Text": {"A": 12, "B": 42, "C": 58, "D": 120},
        "Errors": {"A": 60, "B": 90},
        "How_To_Read": {"A": 24, "B": 100},
    }

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                cell.fill = light_fill

        # Highlight new rows.
        header = [c.value for c in ws[1]]
        if "Is_New" in header:
            col_idx = header.index("Is_New") + 1
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row, col_idx).value).strip().lower() == "yes":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row, col).fill = new_fill
        if "Is_New_Group" in header:
            col_idx = header.index("Is_New_Group") + 1
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row, col_idx).value).strip().lower() == "yes":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row, col).fill = new_fill

        for col, width in widths.get(ws.title, {}).items():
            ws.column_dimensions[col].width = width

    wb.save(path)


def run(args: argparse.Namespace) -> int:
    out_dir = Path(args.out).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(KST).strftime("%Y%m%d_%H%M%S")

    print("FUT.GG RAW objective crawler started.")
    print(f"Output folder: {out_dir}")
    print(f"Collection time KST: {datetime.now(KST).strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Deadline base KST: {scheduled_deadline_base_kst().strftime('%Y-%m-%d %H:%M:%S')}")

    previous_file = find_previous_file(out_dir, args.previous)
    if previous_file:
        print(f"Previous file for Is_New comparison: {previous_file}")
    else:
        print("Previous file not found. Is_New will be No for all objectives.")
    prev_keys = previous_keys_from_file(previous_file)

    print("Collecting objective URLs...")
    links = collect_objective_links(headless=args.headless)
    print(f"Collected active objective URLs: {len(links)}")

    groups: List[Dict[str, Any]] = []
    objectives: List[Dict[str, Any]] = []
    raw_pages: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless)
        context = browser.new_context(viewport={"width": 1450, "height": 1200})
        for idx, link in enumerate(links, start=1):
            print(f"[{idx}/{len(links)}] {link['URL']}")
            try:
                group, objs, raw_text = scrape_objective_page(context, link, idx)
                groups.append(group)
                objectives.extend(objs)
                raw_pages.append({
                    "Group_ID": group["Group_ID"],
                    "Mission_Group_Name": group["Mission_Group_Name"],
                    "URL": group["URL"],
                    "Raw_Page_Text": raw_text,
                })
                if not objs:
                    errors.append({"URL": link["URL"], "Error": "No objective text extracted"})
            except Exception as exc:
                errors.append({"URL": link["URL"], "Error": f"{type(exc).__name__}: {exc}"})
                print(f"  ERROR: {type(exc).__name__}: {exc}")
        browser.close()

    df_groups = pd.DataFrame(groups)
    df_objectives = pd.DataFrame(objectives)
    df_links = pd.DataFrame(links)
    df_raw_pages = pd.DataFrame(raw_pages)

    if df_groups.empty:
        df_groups = pd.DataFrame(columns=["Group_ID", "Mission_Group_Name", "Objective_Type", "URL", "Campaign_Expires", "Campaign_Deadline_KST", "Campaign_Deadline_ISO", "Page_Text_Raw_Short"])
    if df_objectives.empty:
        df_objectives = pd.DataFrame(columns=[
            "Group_ID", "Mission_Group_Name", "Objective_Type", "URL", "Objective_Order", "Objective_Name_Raw", "Objective_Text_Raw", "Reward_Text_Raw", "Mode_Text_Raw", "Difficulty_Text_Raw", "Campaign_Expires", "Campaign_Deadline_KST", "Campaign_Deadline_ISO", "Extraction_Method", "Objective_Key"
        ])

    df_groups, df_objectives = apply_new_flags(df_groups, df_objectives, prev_keys)
    df_manifest = build_manifest(df_groups, df_objectives, errors)

    xlsx_path, json_path = write_outputs(out_dir, df_manifest, df_groups, df_objectives, df_links, df_raw_pages, errors, timestamp)

    print(f"Done XLSX: {xlsx_path}")
    print(f"Done JSON: {json_path}")
    print(f"Latest XLSX copy: {out_dir / 'objectives_raw_latest.xlsx'}")
    print(f"Latest JSON copy: {out_dir / 'objectives_raw_latest.json'}")
    print(f"Groups: {len(df_groups)} | Objectives: {len(df_objectives)} | New objectives: {int((df_objectives.get('Is_New', pd.Series(dtype=str)) == 'Yes').sum()) if not df_objectives.empty else 0} | Errors: {len(errors)}")
    return 0


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="FUT.GG raw objectives crawler")
    parser.add_argument("--out", default="outputs", help="Output directory")
    parser.add_argument("--previous", default="", help="Optional previous raw/latest xlsx for Is_New comparison")
    parser.add_argument("--headless", action="store_true", default=False, help="Run browser headless")
    parser.add_argument("--show", action="store_true", default=False, help="Show browser window")
    args = parser.parse_args(argv)
    if args.show:
        args.headless = False
    return args


if __name__ == "__main__":
    raise SystemExit(run(parse_args()))
